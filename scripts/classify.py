# -*- coding: utf-8 -*-
"""`gw classify` — 키워드 1차 분류 + LLM 처리 큐 생성.

docs/CONTRACT.md 7절의 하이브리드 분류를 구현한다. **LLM 을 호출하지 않는다.**
규칙으로 확신할 수 있는 문항만 자동으로 채우고, 애매한 것만 골라 LLM 이 한 번에
읽을 큐(classify_queue.json)를 만든다. 이 모듈이 토큰 절약의 핵심이다 — 큐가
작을수록, 큐 안의 문항당 정보가 적을수록 LLM 이 쓰는 토큰이 준다.

## 원본과 점수 계산이 다른 이유
원본(D:/codex_work/programs/CSAT_Clipper/core/keyword_classifier.py)은
점수 = 매칭된 키워드 수 / 그 단원에 등록된 키워드 총수 였다. 이러면 키워드를
성실하게 많이 적어둔 단원일수록 분모가 커져 손해를 본다 — 정성껏 80개를 적은
단원이 대충 5개만 적은 단원보다 낮은 점수를 받는 역설이 생긴다.

여기서는 "이 단원에 등록된 키워드 중 몇 개나 맞았나" 대신 "매칭된 키워드들이
이 단원이라는 가설을 얼마나 뒷받침하는가"를 noisy-OR 로 합성한다:
    score = 1 - Π(1 - w_i)   (w_i = 매칭된 키워드 i의 가중치, 0~1)
매칭되지 '않은' 키워드는 이 식에 아예 등장하지 않으므로, 단원에 키워드를 많이
적어둬도 손해가 없다 — 오히려 매칭 기회만 늘어난다.

가중치 w_i 는 두 요소의 곱이다.
1. **역빈도(rarity)**: 문항 코퍼스 전체에서 드물게 등장하는 키워드일수록 그
   키워드가 나온 문항이 특정 단원일 확률에 대한 정보량이 크다. 표준 스무딩
   IDF `log((N+1)/(df+1))+1` 를 [0,1] 로 정규화해서 쓴다.
2. **위치 가중치**: 아래 POSITION_WEIGHT 주석 참조. 해설 > <보기> > 발문 > 선택지
   순으로 신뢰한다(미션 지시: "해설지 텍스트가 있으면 그쪽을 더 신뢰해라").
   같은 키워드가 발문과 보기에 동시에 나오는 것은 흔한 중복 서술이라 이중으로
   세지 않고, 그 키워드가 나온 위치 중 가장 신뢰도 높은 위치의 가중치만 쓴다.

## ★ 반드시 기억할 교훈 — 동사가 아니라 명사로 판정하라
2015→2022 성취기준 매핑 작업에서, 성취기준 문장을 동사(이해한다/설명한다/
토론한다) 위주로 읽어 판정한 결과 **행성우주과학 15개 성취기준 중 9개가
"0문항"으로 잘못 나왔다.** 예를 들어 '탐사의 *성과*를 이해한다'를 "우주 탐사
과정 자체를 다루는 문항만 해당"으로 좁게 읽으면, 허블 법칙으로 적색편이를
계산하는 문항(성과를 '활용'하는 문항)을 전부 놓친다. 130문항을 원문과
재대조해서야 9개 중 3개만 진짜 공백임을 확인했다 — 나머지는 소재(대상=명사)가
같으면 동사가 안 맞아도 '부분 대응'으로 잡아야 했다.

이 교훈은 이 파일의 판정 로직(명사 성격의 키워드 매칭만 하고 동사는 걸러야
한다는 설계 방향)과 classify_queue.json 의 guidance 필드 양쪽에 남긴다 — LLM 이
큐를 읽는 순간 같은 함정에 빠지지 않도록.

## ★ 점수는 신뢰도가 아니다 — 보정(calibration) 없이는 자동확정하지 않는다
실측: 지구과학Ⅱ 문항 80개(학습에 쓰지 않은 4회차)를 교육과정 초안 사전으로 채점했더니
argmax **정확도가 40%** 였고, 자동확정 구간에서도 최고 67% 였다. 그런데 그 오답들에
0.87~0.94 짜리 점수가 붙어 있었다. 점수가
높다는 것과 맞다는 것 사이에 아무 관계가 없었다. 원인은 사전이다 — `standards
--draft-keywords` 가 교육과정 성취기준 **문장**에서 기계로 뽑은 초안이라, 같은
단원의 형제 성취기준들이 '원시 지구의 형성', '미행성체' 같은 명사를 그대로 공유한다
(실측: 지Ⅱ 는 서로 다른 용어 162개 중 43개가 2개 이상의 코드에 동시 등록돼 있다).
공유된 용어는 시험지
코퍼스에서 드물어 IDF 가 높고, 그래서 **엉뚱한 형제 코드에 높은 점수를 만든다.**

그래서 이 모듈은 두 가지를 한다.
1. `--calibrate` : 라벨을 정답지로 놓고 스스로 채점해 정확도-자동확정률 곡선을
   그리고, 목표 정확도를 지키는 가장 낮은 임계값을 권장한다. 그런 임계값이 없으면
   "자동확정 불가"라고 정직하게 적는다. 결과는 calibration.json(기본 자리는 작업공간).
   정답지로 쓰는 라벨은 `manual`(사람 검수)과 `llm`(큐 판정) 둘 다다.
   ★ manual 만 받던 시절에는 순환이 원리적으로 끊겨 있었다 — `--apply` 가 쓰는 값은
     `llm` 이고 `manual` 을 만드는 경로는 `gw map` + mapping.json(4과목에만 존재)뿐이라,
     새 과목은 무엇을 해도 보정에 도달할 수 없었다. 대신 **표본 구성을 숨기지 않는다**:
     `gold_mix` 에 manual/llm 건수를 적고, manual 이 0건이면 `self_graded: true` 와
     경고를 남긴다. 그 수치는 '정답과의 일치율'이 아니라 'LLM 판정과의 일치율'이다.
2. calibration.json 이 없으면 **자동확정을 아예 하지 않는다.** 전부 큐로 보낸다.
   보정 전 점수를 items 의 confidence 에 싣지 않는다 — CONTRACT 4절상 confidence 는
   숫자|null 이므로 null 을 쓰고, 원점수는 `score` 필드에 따로 둔다.
   보정된 뒤 confidence 에 들어가는 값도 원점수가 아니다. **그 점수대에서 실제로
   맞았던 비율의 95% 신뢰 하한**(Wilson)이다. 점추정을 쓰면 10건 중 10건 맞은
   구간에서 confidence 가 1.0 이 되는데, 그 1.0 은 '틀릴 리 없다'가 아니라
   '아직 틀린 걸 못 봤다'는 뜻이다 — 그 구별이 사라지는 자리가 바로 이 모듈이
   고치려는 사고 지점이라, 정의상 1.0 이 될 수 없는 하한을 싣는다.

## 데이터에서 배우는 사전 — `--learn`
교육과정 문장의 어휘와 실제 문항의 어휘는 다르다. 그래서 **이미 정답이 붙은
문항**(classification.<rev>.by 가 manual 또는 llm)에서 변별력 있는 용어를 직접 캔다.

- 후보 생성: 형태소 분석기를 쓰지 않는다(설치 부담). 공백 토큰화 → 조사 꼬리 제거 →
  유니그램 + 인접 바이그램. `PARTICLE_SUFFIXES` 주석 참조.
- 선별: **로그 오즈비**(Monroe et al. 2008 "Fightin' Words" 의 스무딩된 오즈비 계열).
      lor = log((a+α)/(n_c-a+α)) - log((b+α)/(N-n_c-b+α))
  a = 그 성취기준 문항 중 이 용어가 나온 수, b = 나머지 문항 중 나온 수.
  TF-IDF 나 카이제곱 대신 이걸 고른 이유: (1) 필요한 것은 "이 용어가 드문가"가 아니라
  "이 용어가 **다른 성취기준과 이 성취기준을 갈라주는가**"이고, 로그 오즈비가 바로 그
  대비를 재는 통계량이다. TF-IDF 는 드묾만 재기 때문에 형제 코드가 공유하는 희귀어를
  걸러내지 못한다 — 지금 사고가 난 그 지점이다. (2) 클래스당 문항이 3~5개뿐인 희소
  상황에서 카이제곱은 기대빈도 가정이 깨져 못 쓰지만, α 스무딩이 들어간 오즈비는 a=2
  같은 작은 수에서도 폭발하지 않는다. (3) 같은 단원의 형제 성취기준이 자동으로
  '나머지(rest)'에 들어가므로, **단원은 맞고 성취기준만 틀리는** 주 오류 유형을
  정면으로 겨냥한다.
- 가중치: w = sigmoid(lor) × a/(a+k). 뒤의 축소항이 없으면 a=2 짜리 우연을 0.99 로
  믿어버린다 — 위에서 고발한 "틀린 답에 0.96" 을 학습 쪽에서 반복하는 꼴이 된다.
- 병합: 교육과정 유래 키워드는 `curriculum`, 배운 것은 `learned` 로 **구분해서** 남긴다.
  파일의 형태(평면형·구조형·개정형)를 읽고 쓰는 일은 전부 `scripts/keywordsio.py` 가 한다.
  이 모듈은 개정 한 겹을 이미 고른 `{코드: 칸}` 만 다룬다.

## ★ 사전은 개정별로 나뉘어 있다 — 접두사로는 못 가르는 과목이 있다
사회탐구 5개 과목(경제 `12경제`, 윤리와 사상 `12윤사`, 사회·문화 `12사문`,
세계지리 `12세지`, 세계사 `12세사`)은 2015 개정과 2022 개정의 성취기준 접두사가 **같다.**
예전 사전은 코드만으로 키를 잡았고 이 모듈은 접두사로 개정을 갈랐으므로, 이 과목들에서는
두 교육과정이 한 칸에 겹쳐 앉았다 — 2015 문항을 2022 키워드로 채점하는 결과가 나온다.
지금은 `keywords.json` 이 개정 층을 따로 들고 있고 `Subject.keywords(rev)` 가 개정을
**반드시** 인자로 받는다. 접두사 필터는 그 위에서 검산으로만 쓴다.

**이 도구는 쓸수록 정확해진다.** 새 과목도 큐 판정 30~50문항을 `--apply` 한 뒤
`--learn` 을 돌리면 사전이 그 과목의 실제 어휘로 갈아탄다. 교육과정 초안은 출발점일 뿐이다.
다만 **적은 라벨로는 손해일 수 있다.** 실측(한국지리, 라벨 20건 = 한 회차):
성취기준이 18개로 흩어져 문항 2건 이상이 모인 코드가 2개뿐이었고, 용어 17개가
그 2개에만 붙어 홀드아웃 정확도가 0.10 → 0.05 로 내려갔다(top-3 은 0.10 → 0.20).
그래서 `--learn` 은 사전이 몇 개 코드에만 붙으면 경고하고, 이득 여부는
`--calibrate` 를 홀드아웃 회차로 돌려 확인하라고 말한다.

## 산출물은 저장소가 아니라 작업공간에 쓴다
`keywords.json` · `calibration.json` 의 기본 자리는 `<workspace>/` 다. 읽을 때는
작업공간 사본이 저장소 사본을 이기고, `--install` 을 준 실행만 `subjects/<slug>/` 에
쓴다. 이유(git pull 이 막히는 실측 사고)는 `scripts/keywordsio.py` 의
'생성물이 앉는 자리' 주석에 있다. 두 사본이 동시에 있으면 리포트에 warn 을 남긴다.

## 점수 정규화를 바꾼 이유 (share × noisy-OR)
학습 사전을 붙이자 문항당 매칭 용어가 늘어 noisy-OR 점수가 거의 전부 1.0 으로
포화했다. 그러면 CONTRACT 7절의 "1·2위 점수 차 < 0.15" 규칙이 무력해진다 —
1위 0.999 / 2위 0.995 는 압도적 우세인데 격차가 0.004 라 전부 큐로 갔다
(실측: 자동확정률 57.5% → 2.5%). 그래서 점수를 다음으로 바꿨다.

    evidence(c) = -Σ log(1 - w_i)            (noisy-OR 의 로그 영역 — 포화하지 않는다)
    score(c)    = evidence(c)/Σevidence(·) × (1 - exp(-evidence(c)))
                  └ 이 코드가 가진 증거의 '비중'   └ 그 증거의 '절대량'

뒤 항은 정확히 옛 noisy-OR 값이므로 이 식은 옛 점수의 확장이다. 앞 항이 없으면
증거가 늘수록 모든 코드가 1.0 에 붙고, 뒤 항이 없으면 약한 용어 하나만 걸린 코드가
비중 1.0 으로 만점을 받는다. 둘 다 필요하다. 스케일이 [0,1] 로 유지되므로 CONTRACT
7절의 0.35/0.15 를 그대로 쓴다.

## 엑셀 왕복
현장 교사는 keywords.json 을 직접 편집하기보다 엑셀이 편하다. `--export-xlsx` /
`--import-xlsx` 로 왕복한다(원본 utils/excel_template.py 의 "단원=열, 키워드=행"
발상을 유지하되, 열 헤더가 자유 텍스트 단원명이 아니라 **개정 + 성취기준 코드**다 —
개정이 없으면 같은 코드를 쓰는 두 교육과정을 되돌릴 때 구분할 수 없다).
옛 판형(1행이 코드)으로 만든 파일도 계속 읽는다.
"""
from __future__ import annotations

import collections
import json
import math
import re
import time
import unicodedata
from pathlib import Path

import keywordsio  # keywords.json 읽기·쓰기는 전부 이 모듈을 통한다
from common import Space, Report, load_subject
from common.ids import split_qid
from common.paths import CURRICULUM_STANDARDS
from common.progress import track

# --- 큐 판정 임계값. docs/CONTRACT.md 7절에서 이미 못박은 값이라 여기서
#     새로 정하지 않는다. 아래 점수 계산을 이 값들이 뜻있게 작동하도록
#     0~1 스케일로 맞춰 설계했다(선택지 하나 강한 매칭 ≈ 0.5~0.7, 여러 개
#     겹치면 0.8+ 로 포화).
TOP_MIN = 0.35
GAP_MIN = 0.15

# 문항 본문 위치별 신뢰도. 실측하며 조정 가능하지만 지금은 다음 근거로 고정했다.
#   해설(explanation) : 단원 소제목이 그대로 나오는 경우가 많다 (미션 지시) → 최고
#   <보기>(boxed)      : 자료 설명 자체가 핵심 개념을 직접 노출하는 경우가 많다
#   발문(stem)         : "다음 자료에 대한 설명으로 옳은 것은?" 류 정형 문구가
#                        섞여 신호 대 잡음비가 낮다 → 기준선(1.0)
#   선택지(choices)    : 오답 매력도를 위해 일부러 다른 단원 개념을 섞어 넣는
#                        경우가 많다(디스트랙터) → 가장 낮게
POSITION_WEIGHT = {"explanation": 1.5, "boxed": 1.2, "stem": 1.0, "choices": 0.8}

DEFAULT_REVISIONS = ("2015", "2022")
DEFAULT_IDF = 1.0          # 코퍼스에서 아예 못 찾은 키워드(이론상 없어야 함)의 안전망
EXCERPT_LIMIT = 400        # CONTRACT 7절: 문항당 본문 400자 상한. LLM 컨텍스트 방어.
QUEUE_CANDIDATES = 3       # CONTRACT 7절: 후보 성취기준 3개 상한.
# 큐 최상위에 싣는 성취기준 목록의 상한. 한 과목의 한 개정은 실측 15~45개다
# (지구과학Ⅱ 37 · 한국지리 28). 이보다 크면 과목 범위가 안 좁혀진 상태라는 뜻이라
# 목록 대신 그 사실을 싣는다 — `_standards_catalog` 주석 참조.
STANDARDS_CATALOG_LIMIT = 120

# ── --learn 하이퍼파라미터 ───────────────────────────────────────────────
# 지구과학Ⅱ 8회차로 28가지 6/2 회차 분할을 전부 돌려 정한 값이다. 아래 범위
# 안에서는 홀드아웃 성취기준 정확도가 0.675~0.75 사이로만 움직였다 — 이 값들에
# 결과가 아슬아슬하게 걸려 있지 않다는 뜻이라 기본값으로 박아둔다.
LEARN_MIN_DF = 2           # 최소 두 문항에 나와야 후보. 1이면 우연을 외운다.
LEARN_MAX_DF_RATIO = 0.4   # 학습 문항의 40% 초과에 나오는 말은 시험지 상투어('그림','옳은')다.
                           # 한국어 불용어 목록을 손으로 적는 대신 빈도로 거른다 — 과목·언어에
                           # 손대지 않는 방식이라 다른 과목에서도 그대로 작동한다.
LEARN_MIN_LOR = 1.5        # 로그 오즈비 하한. 1.5 ≈ 오즈비 4.5배.
LEARN_TOP_PER_CODE = 15    # 코드당 상한. 사람이 눈으로 훑을 수 있는 분량을 유지한다.
LEARN_ALPHA = 0.5          # 오즈비 스무딩(Jeffreys prior). 0 나눗셈과 폭주를 막는다.
LEARN_SHRINK_K = 2.0       # w = sigmoid(lor) × a/(a+k). 근거 문항 수가 적으면 가중치를 깎는다.
LEARN_MIN_ITEMS = 20       # 이보다 적은 라벨로 배우면 사전이 아니라 잡음이다.

# ── --calibrate ─────────────────────────────────────────────────────────
CALIB_TARGET_ACCURACY = 0.85   # 이걸 못 지키면 자동확정 불가로 적는다.
CALIB_THRESHOLDS = tuple(round(0.20 + 0.05 * i, 2) for i in range(15))  # 0.20~0.90
# 자동확정 표본 하한. 처음 5로 뒀다가 실측에서 **5건 중 5건 정답 → 권장 임계값 0.90**
# 이라는 답이 나왔다. 5/5 의 95% 하한은 0.57 이다 — 즉 "정확도 100%" 는 표본이
# 작아서 나온 말이었고, 이것을 권장값으로 내보내면 이 모듈이 고치려던 바로 그 사고
# (작은 근거에 붙은 큰 확신)를 보정 단계에서 되풀이하게 된다.
CALIB_MIN_AUTO = 10
CALIB_MIN_HOLDOUT = 20         # 학습에 쓰지 않은 채점 표본의 하한.
CALIB_Z = 1.96                 # 95% 신뢰구간
CALIBRATION_FILE = "calibration.json"

# ── 엑셀 왕복 판형 ──────────────────────────────────────────────────────
# 1행 개정 / 2행 성취기준 코드 / 3행 단원 라벨 / 4행부터 키워드.
# 1행이 늘어난 이유는 `_export_xlsx` 주석 참조(같은 코드를 두 개정이 쓴다).
XLSX_FIRST_TERM_ROW = 4

_WS_RE = re.compile(r"\s+")


# ---------------------------------------------------------------- 텍스트 유틸
def _normalize(text) -> str:
    if not text:
        return ""
    t = str(text).replace("\n", " ").replace("\r", " ")
    return _WS_RE.sub(" ", t).strip()


def _compact(text: str) -> str:
    """공백만 제거한 버전. 전사자마다 띄어쓰기가 들쭉날쭉해 '해양 판'/'해양판'
    처럼 같은 말이 다르게 적히는 사고가 잦았다 — 공백 버전이 실패할 때만
    보조로 쓴다(완전히 이걸로 대체하면 서로 다른 단어가 우연히 이어붙어
    거짓 매칭이 날 위험이 있다)."""
    return text.replace(" ", "")


def _extract_fields(item: dict) -> dict[str, str]:
    """items/<qid>.json 의 text 블록에서 위치별 텍스트를 뽑는다.

    CONTRACT.md 4절의 items 스키마 예시는 text.{stem,boxed,choices} 까지만
    보여준다 — 해설(explanation) 필드는 아직 계약에 없다. extract 단계가
    해설 텍스트를 어딘가에 남긴다면(text.explanation / text.solution /
    item.solution 등) 여기서 최대한 찾아서 쓰고, 없으면 조용히 건너뛴다.
    이 필드가 정말 필요하면 todo 로 스키마 추가를 요청해야 한다.
    """
    text = item.get("text") or {}
    fields: dict[str, str] = {}
    if text.get("stem"):
        fields["stem"] = text["stem"]
    if text.get("boxed"):
        fields["boxed"] = text["boxed"]
    choices = text.get("choices")
    if choices:
        if isinstance(choices, list):
            fields["choices"] = " / ".join(str(c) for c in choices)
        else:
            fields["choices"] = str(choices)
    explanation = (
        text.get("explanation") or text.get("solution") or text.get("해설")
        or (item.get("solution") if isinstance(item.get("solution"), str) else None)
    )
    if explanation:
        fields["explanation"] = explanation
    return fields


def _build_excerpt(item: dict, limit: int = EXCERPT_LIMIT) -> str:
    fields_raw = item.get("text") or {}
    if not fields_raw and item.get("extraction_mode") == "vision":
        # 텍스트 레이어 없는 회차(예: 2025 수능) — CONTRACT 4절: text 가 비어도
        # 파이프라인은 끝까지 돌아야 한다. LLM 에게 본문 대신 크롭 경로를 준다.
        crop = item.get("crop", "(crop 경로 없음)")
        return f"[vision 모드 — 텍스트 없음. 크롭 이미지 판독 필요: {crop}]"[:limit]

    fields = _extract_fields(item)
    # 신뢰 순서(POSITION_WEIGHT 와 동일한 근거): 해설 > 보기 > 발문 > 선택지.
    order = [("해설", fields.get("explanation")), ("보기", fields.get("boxed")),
             ("발문", fields.get("stem")), ("선택지", fields.get("choices"))]
    out = ""
    for label, chunk in order:
        if not chunk:
            continue
        piece = f"[{label}] {_normalize(chunk)} "
        if len(out) + len(piece) > limit:
            remaining = limit - len(out)
            if remaining > 10:
                out += piece[:remaining]
            break
        out += piece
    return out.strip()[:limit]


# ------------------------------------------------------- keywords.json 스키마
# 파일의 세 형태(평면형·구조형·개정형)를 읽고 쓰는 일은 전부 scripts/keywordsio.py 가
# 한다. 이 모듈은 **개정 한 겹을 이미 고른 상태**의 {코드: 칸} 만 받는다.
#
# 예전에는 여기서 직접 json.load 하고 접두사로 개정을 갈랐다. 그런데 사회탐구
# 5개 과목(경제·윤리와 사상·사회·문화·세계지리·세계사)은 2015 와 2022 의 성취기준
# 접두사가 **같다**(둘 다 12윤사 등). 접두사로는 개정을 가를 수 없고, 파일도 코드만으로
# 키를 잡았으므로 한쪽이 다른 쪽을 조용히 덮어썼다 — 그래서 2015 문항을 2022 키워드로
# 채점하게 된다. 개정을 파일 구조로 올려 그 가능성 자체를 없앴다.


def _note_scope(report, scope) -> None:
    """개정 판정이 접두사 폴백으로 떨어졌으면 리포트에 올린다.

    폴백 자체는 옛 동작이라 결과가 나빠지지는 않지만, 그 상태에서는 접두사가
    같은 두 개정을 못 가른다. 조용히 떨어지면 그 사실을 아무도 모른다.
    """
    if scope.why:
        report.note(f"{scope.revision} 성취기준 범위", scope.why, "warn")


def _keyword_view(entries: dict[str, dict],
                  scope) -> tuple[dict[str, list[str]], dict, dict]:
    """한 개정의 {코드: 칸} → (코드별 용어 목록, (코드,용어)별 학습 가중치, 통계).

    점수 계산기는 "코드마다 용어 목록"만 알면 되므로 여기서 평평하게 만든다.
    학습 가중치는 별도 dict 로 넘긴다 — 학습된 용어는 코퍼스 역빈도(IDF)가 아니라
    라벨에서 직접 잰 변별력을 써야 하기 때문이다(IDF 는 드묾만 재고, 형제 성취기준이
    공유하는 희귀어를 구분하지 못한다. 그게 이 사전이 40% 로 틀린 이유였다).

    scope 로 한 번 더 거르는 이유: 개정 층을 믿되 검산은 한다. 옛 형태 파일을
    역추정해 읽은 경우 그 개정 층에 남의 코드가 섞여 들어올 수 있다.
    ★ 예전엔 이 검산을 `code.startswith(접두사)` 로 했는데, 그 비교는 2015 통합과학
      접두사 `10통과` 로 2022 코드 `10통과1-01-01` 까지 통과시킨다. 그러면 2015
      후보 사전에 2022 성취기준이 섞이고, 그 문항은 **에러 없이** 2022 코드로
      자동확정된다. 이제는 `subject.code_scope(개정)` 이 개정 목록 소속을 본다.
    """
    candidates: dict[str, list[str]] = {}
    weights: dict[tuple[str, str], float] = {}
    n_cur = n_learn = 0
    for code, entry in entries.items():
        if code not in scope:
            continue
        cur = list(entry.get("curriculum") or [])
        terms = list(cur)
        n_cur += len(cur)
        for row in entry.get("learned") or []:
            term = str(row.get("term") or "")
            if not term:
                continue
            if term not in terms:
                terms.append(term)
            n_learn += 1
            w = row.get("weight")
            if isinstance(w, (int, float)) and not isinstance(w, bool):
                weights[(code, term)] = float(w)
        if terms:
            candidates[code] = terms
    stats = {"codes": len(candidates), "curriculum_terms": n_cur, "learned_terms": n_learn}
    return candidates, weights, stats


# ---------------------------------------------------------------- IDF / 점수
def _compute_idf(corpus_texts: list[str], keywords: set[str]) -> dict[str, float]:
    """코퍼스(회차 전체 문항) 기준 역빈도. 개정(2015/2022)별로 다시 계산하지
    않는다 — 어떤 단어가 흔한지 드문지는 시험지 코퍼스의 속성이지, 그 단어를
    어느 성취기준에 매달아뒀는지와는 무관하다."""
    n = len(corpus_texts)
    norm_docs = [(_normalize(t), _compact(_normalize(t))) for t in corpus_texts]
    idf_max = math.log((n + 1) / 1) + 1  # df=0(가장 희귀)일 때의 상한값 — 정규화 분모

    idf: dict[str, float] = {}
    for kw in keywords:
        kw_n = _normalize(kw)
        if not kw_n or kw_n in idf:
            continue
        kw_c = _compact(kw_n)
        df = 0
        for norm_t, compact_t in norm_docs:
            if kw_n in norm_t or (kw_c and kw_c in compact_t):
                df += 1
        raw = math.log((n + 1) / (df + 1)) + 1
        idf[kw_n] = (raw / idf_max) if idf_max > 0 else 1.0
    return idf


def _score_item(fields: dict[str, str], candidates: dict[str, list[str]],
                 idf: dict[str, float],
                 learned_w: dict | None = None) -> tuple[dict[str, float], dict[str, int]]:
    """성취기준별 점수 = 증거 비중 × 증거 절대량. 상세 근거는 모듈 docstring 참조.

    옛 식(순수 noisy-OR)은 매칭 용어가 늘면 모든 코드가 1.0 에 붙어 CONTRACT 7절의
    '1·2위 격차' 규칙을 무력화했다. 여기서는 noisy-OR 을 로그 영역(evidence)에서
    다루고, 그 중 이 코드가 차지하는 비중을 곱한다.
    """
    learned_w = learned_w or {}
    norm_fields = {name: (_normalize(t), _compact(_normalize(t)))
                   for name, t in fields.items() if t}

    evidence: dict[str, float] = {}
    matched: dict[str, int] = {}
    for code, kws in candidates.items():
        prob_no_evidence = 1.0
        hit = 0
        for kw in kws:
            kw_n = _normalize(kw)
            if not kw_n:
                continue
            kw_c = _compact(kw_n)
            best_pos_w = 0.0
            for field_name, (norm_t, compact_t) in norm_fields.items():
                if kw_n in norm_t or (kw_c and kw_c in compact_t):
                    best_pos_w = max(best_pos_w, POSITION_WEIGHT.get(field_name, 1.0))
            if best_pos_w > 0:
                hit += 1
                # 학습된 용어는 라벨에서 직접 잰 변별력을 쓴다. 없으면 코퍼스 역빈도.
                base = learned_w.get((code, kw))
                if base is None:
                    base = idf.get(kw_n, DEFAULT_IDF)
                w = min(0.97, base * best_pos_w)
                prob_no_evidence *= (1 - w)
        # -log(1-noisy_or). 0.97 상한 덕에 0 이 되지 않지만, 방어적으로 하한을 둔다.
        evidence[code] = -math.log(max(prob_no_evidence, 1e-12))
        matched[code] = hit

    total = sum(evidence.values())
    scores: dict[str, float] = {}
    for code, ev in evidence.items():
        share = (ev / total) if total > 0 else 0.0
        magnitude = 1.0 - math.exp(-ev)      # 옛 noisy-OR 점수와 동일한 값
        scores[code] = round(share * magnitude, 4)
    return scores, matched


def _rank(scores: dict[str, float]) -> list[tuple[str, float]]:
    """동점일 때 코드 순으로 안정 정렬. 정렬이 흔들리면 재실행마다 판정이 바뀐다."""
    return sorted(scores.items(), key=lambda kv: (-kv[1], kv[0]))


def _queue_reasons(ranked, matched, top_min: float) -> list[str]:
    """CONTRACT 7절의 큐 편입 사유. 자동확정 임계값만 보정값으로 갈아끼운다."""
    top1 = ranked[0] if ranked else (None, 0.0)
    top2 = ranked[1] if len(ranked) > 1 else (None, 0.0)
    reasons = []
    if sum(matched.values()) == 0:
        reasons.append("no_match")
    if top1[1] < top_min:
        reasons.append("low_top")
    # top1 이 증거 없이(matched=0) 0점으로 동률인 경우까지 "간발의 차"라고 부르면
    # 오해를 준다 — no_match 가 이미 그 상황을 설명한다.
    if top2[0] and matched.get(top1[0], 0) > 0 and (top1[1] - top2[1]) < GAP_MIN:
        reasons.append("close_gap")
    return reasons


# ---------------------------------------------------------------- 성취기준 메타
def _load_standards_meta(revision: str) -> dict[str, dict]:
    """curriculum/standards/<revision>.json 에서 {코드: {unit, title}} 을 읽는다.

    이 파일의 스키마는 아직 docs/CONTRACT.md 에 정의돼 있지 않다(standards
    명령 소관). 없어도 classify 는 끝까지 돌아야 하므로, 없으면 code 만 기록하고
    unit 라벨은 비운다.

    ── 왜 재귀로 훑게 바뀌었나 ──
    처음엔 스키마를 `[{"code","unit","title"}]` 또는 `{코드: {...}}` 로 **추측**했다.
    통합 검증에서 실제 파일을 붙여 보니 실제 모양은 3단 중첩이었다:
        {"revision", "sources", "subjects": {과목명: {"units":
            [{"no", "title", "standards": [{"code", "text", "page"}]}]}}}
    이 추측은 예외를 던지지 않고 **조용히 실패**해서 더 나빴다 — dict 분기가
    최상위 키 'revision'/'sources'/'subjects' 를 성취기준 코드로 착각해
    meta 를 쓰레기로 채우고, 결국 모든 unit 라벨이 None 이 됐는데 아무도
    모른다. 그래서 층 이름을 믿지 않고 트리 전체에서 코드 레코드를 줍고,
    단원 제목(그 레코드를 감싼 상위 노드의 "title")을 함께 물고 내려간다.
    """
    path = CURRICULUM_STANDARDS / f"{revision}.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    meta: dict[str, dict] = {}
    _walk_standards(data, None, meta)
    return meta


def _walk_standards(node, unit_title: str | None, meta: dict[str, dict]) -> None:
    """{"code": ...} 레코드를 줍되, 가장 가까운 상위 노드의 "title"(=단원명)을 물고 내려간다."""
    if isinstance(node, dict):
        code = node.get("code")
        if isinstance(code, str):
            # 레코드 자신이 unit 을 들고 있으면 그쪽이 우선(옛 평면 형태 호환).
            # "title" 은 단원명 자리다. 성취기준 **문장**은 "text" 에 따로 둔다 —
            # _unit_label 의 `unit or title` 폴백에 문장이 새어 들어가면
            # 사이드바 단원명이 통째로 한 문장이 돼 버린다.
            meta.setdefault(code, {
                "unit": node.get("unit") or unit_title,
                "title": unit_title,
                "text": node.get("text"),
            })
            return
        here = _unit_text(node) or unit_title
        for v in node.values():
            _walk_standards(v, here, meta)
    elif isinstance(node, list):
        for v in node:
            _walk_standards(v, unit_title, meta)


def _unit_text(node: dict) -> str | None:
    """단원 노드({"no": 3, "title": "한반도의 지질"})를 "(3) 한반도의 지질" 로 만든다.

    번호를 굳이 붙이는 이유: `gw map` 이 subjects/<slug>/mapping.json 에서 실어 오는
    단원 라벨이 이미 "(3) 한반도의 지질" 형식이다. build.py 는 단원 **문자열**을 키로
    트리를 묶으므로, classify 가 번호 없는 "한반도의 지질" 을 쓰면 같은 단원이 사이드바에
    두 칸으로 갈라진다. 라벨 형식을 여기서 맞춰 둔다.
    """
    title = node.get("title")
    if not isinstance(title, str) or not title.strip():
        return None
    no = node.get("no")
    return f"({no}) {title}" if isinstance(no, int) else title


def _unit_label(meta_by_rev: dict[str, dict], rev: str, code: str) -> str | None:
    row = meta_by_rev.get(rev, {}).get(code) or {}
    return row.get("unit") or row.get("title")


def _standard_text(meta_by_rev: dict[str, dict], rev: str, code: str) -> str | None:
    """성취기준 문장. 없으면 None — 지어내지 않는다."""
    row = meta_by_rev.get(rev, {}).get(code) or {}
    text = row.get("text")
    return _normalize(text) if text else None


def _standards_catalog(subject, meta_by_rev: dict[str, dict], rev: str) -> list[dict]:
    """이 과목·이 개정의 성취기준 전체 목록. 큐 최상위에 한 번만 실린다.

    범위 판정은 `subject.code_scope(rev)` 한 곳에서만 한다(CONTRACT 3절) —
    접두사 문자열 비교로 거르면 통합과목처럼 접두사가 겹치는 자리에서 남의 개정
    코드가 목록에 섞이고, 큐를 읽은 LLM 이 그 코드를 골라 `--apply` 로 들어온다.

    상한을 두는 이유: 아직 `curriculum`·`standard_prefixes` 를 안 채운 과목은
    scope 의 basis 가 `any` 라 **개정 파일 전체**(2022 는 2,537개)를 통과시킨다.
    그대로 실으면 큐 하나가 교육과정 전서가 된다 — 토큰 방어가 이 모듈의 존재
    이유다(CONTRACT 7절). 넘치면 싣지 않고 왜 안 실었는지만 남긴다.
    """
    scope = subject.code_scope(rev)
    codes = sorted(scope.filter(meta_by_rev.get(rev, {})))
    if len(codes) > STANDARDS_CATALOG_LIMIT:
        return [{"_omitted": len(codes),
                 "why": f"이 과목의 {rev} 성취기준 범위가 {len(codes)}개로 상한"
                        f"({STANDARDS_CATALOG_LIMIT})을 넘어 목록을 싣지 않았다 — "
                        f"subject.json 의 curriculum.{rev}/standard_prefixes.{rev} 가 "
                        f"비어 범위가 좁혀지지 않은 상태다. 후보 3개로 판정하거나 "
                        f"curriculum/standards/{rev}.json 을 직접 봐라"}]
    return [{"code": c, "unit": _unit_label(meta_by_rev, rev, c),
             "text": _standard_text(meta_by_rev, rev, c)} for c in codes]


# ── 큐에 실리는 지침 ────────────────────────────────────────────────────
# ★ 과목 이름을 쓰지 않는다. 예전 문구에는 '행성우주과학'·'허블 법칙'·'적색편이'
#   가 그대로 박혀 있어서, 한국지리 큐를 읽는 LLM 이 지구과학 사례를 판정 지침으로
#   받았다. 레퍼런스 구현이 문자열로 새어 나온 자리다(CONTRACT 0절: 과목별 차이는
#   데이터로만). 교훈 자체는 과목과 무관하므로 사례를 지우고 규칙만 남긴다.
QUEUE_GUIDANCE = (
    "성취기준은 학습 활동의 동사(이해한다/설명한다/토론한다/조사한다)가 아니라 "
    "그 성취기준이 다루는 **대상(명사)**으로 판정하라. 문항이 그 대상을 소재로 쓰면, "
    "동사가 안 맞아도 부분 대응으로 잡는다. 동사 위주로 좁게 읽어 한 단원의 성취기준 "
    "여럿이 '해당 문항 0건'으로 잘못 나오고 100문항 넘게 재검토한 사고가 실제로 있었다. "
    "candidates 는 키워드 점수 상위 3개일 뿐이다 — 맞는 것이 없으면 standards 목록에서 "
    "직접 고르고, 그래도 하나로 못 정하겠으면 그 문항은 results 에서 **빼라**. "
    "빈 채로 두는 편이 틀린 코드를 박는 것보다 언제나 낫다(빠진 문항은 큐에 남는다)."
)

# ── classify_result.json 의 모양 ────────────────────────────────────────
# 큐 파일이 이 스키마를 함께 싣는다. 이유는 queue_payload 주석 참조.
RESULT_SCHEMA = {
    "file": "classify_result.json (이름은 자유. --apply 에 그 경로를 준다)",
    "shape": "최상위는 배열이거나 {\"results\": [...]} 객체. 둘 다 받는다.",
    "fields": {
        "qid": "필수. 큐 항목의 qid 를 그대로.",
        "revision": "필수. 큐 항목의 revision 을 그대로(\"2015\"/\"2022\").",
        "standard": "필수. 성취기준 코드. candidates 나 standards 목록에 있는 코드여야 한다 — "
                    "목록에 없는 코드는 errors 로 거부된다.",
        "unit": "선택. 비우면 curriculum/standards 에서 코드로 채운다. "
                "직접 적을 거면 standards 목록의 unit 문자열을 글자 그대로 써라 — "
                "다르게 적으면 제작기 사이드바에서 같은 단원이 두 칸으로 갈린다.",
        "notes": "선택. 판정 근거·망설인 이유. items 의 notes 에 남는다.",
        "confidence": "선택. 적어도 items 의 confidence 에는 실리지 않는다 — "
                      "그 칸은 보정으로 실측한 정확도의 95% 하한 자리다(CONTRACT 4절). "
                      "네가 적은 값은 ext.llm_confidence 로만 남는다.",
    },
    "example": {"results": [
        {"qid": "<큐의 qid>", "revision": "2015", "standard": "<큐의 candidates[0].code>",
         "notes": "발문의 대상이 이 성취기준의 소재와 같다"},
    ]},
    "apply": "python scripts/gw.py classify --subject <slug> --apply <이 파일 경로> "
             "[--workspace <작업공간>]",
}


# ------------------------------------------------------------ 공용: 문항 적재
def _load_items(space: Space, report: Report) -> dict[str, dict]:
    items: dict[str, dict] = {}
    for p in space.iter_items():
        try:
            items[p.stem] = json.loads(p.read_text(encoding="utf-8"))
        except Exception as exc:
            report.note(p.stem, f"items json 파싱 실패: {exc}", "error")
    return items


def _item_text(item: dict) -> str:
    return " ".join(_extract_fields(item).values())


def _selection(raw: str | None) -> set[str] | None:
    """--only 값 → 토큰 집합. qid 와 exam_id 를 모두 받는다(형제 모듈과 같은 규약)."""
    if not raw:
        return None
    return {t.strip() for t in raw.split(",") if t.strip()}


def _is_qid(token: str) -> bool:
    try:
        split_qid(token)
        return True
    except ValueError:
        return False


def _selected(qid: str, only: set[str] | None) -> bool:
    if not only:
        return True
    if qid in only:
        return True
    try:
        return split_qid(qid)[0] in only
    except ValueError:
        return False


def _gold_labels(subject, items: dict[str, dict], rev: str,
                 accept_by: tuple[str, ...]) -> tuple[dict[str, str], dict[str, str]]:
    """정답지 역할을 할 라벨을 모은다. → ({qid: 성취기준코드}, {qid: 라벨 출처})

    두 곳을 본다.
      1) items/<qid>.json 의 classification.<rev> 중 by 가 accept_by 에 든 것
      2) subjects/<slug>/mapping.json 의 같은 문항 (by 가 accept_by 에 든 것)
    mapping.json 을 함께 보는 이유: `gw map` 을 아직 안 돌린 작업공간에서도 채점이
    되어야 하기 때문이다. 사람이 만든 매핑 파일은 이미 저장소에 커밋돼 있다.

    ★ 두 번째 반환값을 집계(Counter)가 아니라 **qid별 출처**로 바꾼 이유:
      부르는 쪽이 `--only` 로 표본을 자른 뒤 다시 세야 하는데, 예전에는 items 만
      되짚어 세느라 mapping.json 에서 온 라벨이 집계에서 통째로 사라졌다.
      그리고 아래 `_calibrate` 가 **채점 표본에 사람 라벨이 몇 건, LLM 라벨이 몇
      건인지**를 파일과 리포트에 그대로 실어야 한다 — 그 수치 없이는 정확도를
      얼마나 믿을지 읽는 쪽이 판단할 수 없다.
    """
    gold: dict[str, str] = {}
    source: dict[str, str] = {}
    for qid, item in items.items():
        entry = (item.get("classification") or {}).get(rev) or {}
        if entry.get("by") in accept_by and entry.get("standard"):
            gold[qid] = entry["standard"]
            source[qid] = str(entry["by"])

    mapping = subject.mapping() or {}
    for qid, entry in (mapping.get("items") or {}).items():
        if qid in gold or qid not in items:
            continue
        row = (entry or {}).get(rev) or {}
        if row.get("by") in accept_by and row.get("standard"):
            gold[qid] = row["standard"]
            # 출처를 파일 단위로 구분해 둔다. mapping.json 은 사람이 만든
            # 매핑이라 by=manual 이지만, 어느 파일에서 왔는지는 남겨야
            # "이 정확도를 뭘로 쟀나"를 되짚을 수 있다.
            source[qid] = f"mapping.json({row.get('by')})"
    return gold, source


def _label_mix(source: dict[str, str], qids) -> dict[str, int]:
    """{라벨 출처: 건수}. 사람 라벨과 LLM 라벨의 비율을 리포트가 그대로 싣는다."""
    counted = collections.Counter(source.get(q, "unknown") for q in qids)
    return {k: counted[k] for k in sorted(counted)}


def _human_share(source: dict[str, str], qids) -> tuple[int, int]:
    """(사람 라벨 수, LLM 라벨 수). 'manual' 이 들어간 출처만 사람으로 센다."""
    human = sum(1 for q in qids if "manual" in source.get(q, ""))
    return human, len(list(qids)) - human


# ------------------------------------------------------------ 공용: 용어 캐기
# 조사 꼬리. 형태소 분석기 없이 '지진파가' / '지진파는' / '지진파의' 를 한 낱말로
# 모으기 위한 최소 장치다. 긴 것부터 검사해야 '에서는' 이 '는' 으로 먼저 잘리지 않는다.
# 어간이 2자 미만으로 줄어들면 자르지 않는다 — '바다' 의 '다' 같은 오절단 방지.
PARTICLE_SUFFIXES = ("에서는", "으로는", "에게서", "이라는", "라는", "에서", "에게",
                     "으로", "까지", "부터", "보다", "처럼", "마다", "조차", "라도",
                     "이나", "와의", "과의", "의", "은", "는", "이", "가", "을", "를",
                     "와", "과", "도", "만", "로", "에", "나", "란", "며", "고")
_TOKEN_RE = re.compile(r"[가-힣A-Za-z0-9]+")


def _strip_particle(token: str) -> str:
    for suf in PARTICLE_SUFFIXES:
        if len(token) > len(suf) + 1 and token.endswith(suf):
            return token[: -len(suf)]
    return token


def _terms_of(text: str) -> set[str]:
    """문항 본문 → 후보 용어 집합(유니그램 + 인접 바이그램).

    NFKC 정규화는 PITFALLS 3-5 의 지시다(로마숫자·전각). 한 문항 안에서 같은 말이
    여러 번 나와도 한 번으로 센다 — 로그 오즈비는 '몇 문항에 나왔나'(문서빈도)를
    쓰는 통계라 문항 내 반복은 정보가 아니라 잡음이다.
    """
    norm = unicodedata.normalize("NFKC", text or "")
    tokens = [_strip_particle(t) for t in _TOKEN_RE.findall(norm)]
    tokens = [t for t in tokens if len(t) >= 2 and not t.isdigit()]
    terms = set(tokens)
    terms.update(f"{a} {b}" for a, b in zip(tokens, tokens[1:]))
    return terms


def _mine_terms(docs: dict[str, set[str]],
                labels: dict[str, str]) -> tuple[dict[str, list[dict]], dict]:
    """라벨된 문항에서 성취기준별 변별 용어를 캔다. → (사전, 깔때기 진단)

    두 번째 반환값이 있는 이유: **아무것도 못 배웠을 때 왜 못 배웠는지**를
    리포트가 말할 수 있어야 한다. 실측 사고 — 한국지리 라벨 20건을 주고
    `--learn` 을 돌렸더니 `learned_terms=0` 으로 조용히 끝났다. 리포트에는
    "배운 것이 없다"만 있었고, 사용자는 라벨이 모자란 건지·본문이 없는 건지·
    임계값에 걸린 건지 알 방법이 없었다. 실제 원인은 **라벨이 성취기준마다
    1건씩 흩어져 있어서** `a >= LEARN_MIN_DF(2)` 를 만족하는 코드가 거의 없었던
    것인데, 그건 사용자가 회차를 하나 더 판정하면 바로 풀리는 문제다.
    어느 관문에서 몇 개가 떨어졌는지 세어 돌려준다.
    """
    n_docs = len(docs)
    df = collections.Counter()
    for terms in docs.values():
        df.update(terms)

    by_code: dict[str, list[str]] = collections.defaultdict(list)
    for qid, code in labels.items():
        if qid in docs:
            by_code[code].append(qid)

    diag = {
        "labeled_docs": n_docs,
        "codes": len(by_code),
        "max_items_per_code": max((len(v) for v in by_code.values()), default=0),
        "codes_meeting_min_df": sum(1 for v in by_code.values() if len(v) >= LEARN_MIN_DF),
        "candidate_terms": 0,     # 코드×용어 쌍으로 검사한 총 후보 수
        "dropped_rare": 0,        # 그 성취기준 문항 중 1건에만 나온 용어 (min_df 미달)
        "dropped_common": 0,      # 전체 라벨의 max_df_ratio 를 넘게 나온 상투어
        "dropped_low_lor": 0,     # 변별력(로그 오즈비) 하한 미달
        "dropped_substring": 0,   # 더 긴 용어에 흡수됨
        "over_cap": 0,            # 코드당 상한(top_per_code) 초과
        "kept": 0,
    }

    learned: dict[str, list[dict]] = {}
    for code, qids in by_code.items():
        n_c = len(qids)
        in_class = collections.Counter()
        for qid in qids:
            in_class.update(docs[qid])

        rows = []
        for term, a in in_class.items():
            diag["candidate_terms"] += 1
            if a < LEARN_MIN_DF:
                diag["dropped_rare"] += 1
                continue
            if df[term] > LEARN_MAX_DF_RATIO * n_docs:
                diag["dropped_common"] += 1
                continue
            b = df[term] - a
            lor = (math.log((a + LEARN_ALPHA) / (n_c - a + LEARN_ALPHA))
                   - math.log((b + LEARN_ALPHA) / (n_docs - n_c - b + LEARN_ALPHA)))
            if lor < LEARN_MIN_LOR:
                diag["dropped_low_lor"] += 1
                continue
            weight = (1.0 / (1.0 + math.exp(-lor))) * (a / (a + LEARN_SHRINK_K))
            rows.append({"term": term, "weight": round(weight, 4),
                          "df": a, "lor": round(lor, 3)})
        rows.sort(key=lambda r: (-r["weight"], -r["df"], r["term"]))

        # 부분문자열 흡수: '해양판 섭입' 이 뽑혔으면 그 안의 '해양판' 은 같은 증거를
        # 두 번 세게 만든다(noisy-OR 은 독립을 가정한다). 긴 쪽만 남긴다.
        kept: list[dict] = []
        for row in rows:
            if any(row["term"] != k["term"] and row["term"] in k["term"] for k in kept):
                diag["dropped_substring"] += 1
                continue
            if len(kept) >= LEARN_TOP_PER_CODE:
                diag["over_cap"] += 1
                continue
            kept.append(row)
        if kept:
            learned[code] = kept
            diag["kept"] += len(kept)
    return learned, diag


def _learn_why_zero(diag: dict) -> str:
    """깔때기 진단 → 사람이 읽는 한 문장. **왜 0인지**를 먼저 말한다.

    리포트만 읽는 쪽(CONTRACT 2절)이 다음 행동을 정할 수 있어야 하므로,
    원인마다 '그래서 뭘 하면 되는지'를 붙인다.
    """
    if not diag.get("labeled_docs"):
        return ("학습에 쓸 본문이 한 건도 없다 — 라벨은 있으나 items 의 text 가 비어 있다"
                "(vision 회차이거나 extract 가 아직 안 돌았다)")
    if diag.get("codes_meeting_min_df", 0) == 0:
        return (f"라벨 {diag['labeled_docs']}건이 성취기준 {diag['codes']}개에 흩어져 있다 — "
                f"한 성취기준에 모인 문항이 최대 {diag['max_items_per_code']}건이라 "
                f"최소 문서빈도 {LEARN_MIN_DF}건을 넘는 성취기준이 하나도 없다. "
                f"용어는 '같은 성취기준 문항 {LEARN_MIN_DF}건 이상에 나온 말'만 뽑는다"
                f"(1건짜리는 우연이라). **회차를 하나 더 판정해 --apply 하면 열린다** — "
                f"같은 성취기준이 회차마다 다시 출제되기 때문이다")
    parts = []
    if diag.get("dropped_rare"):
        parts.append(f"그 성취기준 문항 1건에만 나온 용어 {diag['dropped_rare']}개")
    if diag.get("dropped_common"):
        parts.append(f"라벨의 {int(LEARN_MAX_DF_RATIO * 100)}% 넘게 나오는 상투어 "
                     f"{diag['dropped_common']}개")
    if diag.get("dropped_low_lor"):
        parts.append(f"변별력(로그 오즈비 {LEARN_MIN_LOR}) 미달 {diag['dropped_low_lor']}개")
    return (f"후보 용어 {diag.get('candidate_terms', 0)}개가 전부 걸러졌다 — "
            + ", ".join(parts) + ". 라벨을 늘리면 같은 용어가 여러 문항에 걸리며 살아난다"
            if parts else "후보 용어가 없다")


# ------------------------------------------------------------ 공용: 보정 파일
def _wilson_lower(hits: int, n: int, z: float = CALIB_Z) -> float | None:
    """정확도의 95% 신뢰 하한(Wilson score interval).

    점추정(맞은 수/전체)을 그대로 confidence 로 쓰면 10건 중 10건이 맞았을 때
    1.0 이 된다. 그 1.0 은 '틀릴 리 없다'가 아니라 '아직 틀린 걸 못 봤다'는 뜻인데,
    items 에 실리는 순간 아무도 그 구별을 하지 않는다 — 이 모듈이 고치려는 사고가
    정확히 그것이었다. 그래서 items 의 confidence 에는 **하한**을 싣는다.
    하한은 정의상 1.0 이 될 수 없고, 표본이 늘면 점추정에 수렴한다.
    Wilson 을 쓰는 이유: 정규근사(Wald)는 p̂=1.0 에서 구간 폭이 0 이 되어
    "10/10 → 하한 1.0" 이라는 같은 거짓말을 되풀이한다.
    """
    if n <= 0:
        return None
    p = hits / n
    denom = 1 + z * z / n
    centre = p + z * z / (2 * n)
    margin = z * math.sqrt(max(p * (1 - p) / n + z * z / (4 * n * n), 0.0))
    return round(max(0.0, (centre - margin) / denom), 4)


def _calibration_repo_path(subject) -> Path:
    # Subject 에 calibration 경로 속성이 없다(common/ 은 다른 사람 소관이라 안 건드린다).
    # keywords.json 과 같은 폴더 규약을 그대로 따른다.
    return subject.keywords_path.parent / CALIBRATION_FILE


def _load_calibration(subject, space, report=None) -> dict:
    """보정 파일을 읽는다. 작업공간 사본이 저장소 사본을 이긴다(keywordsio 주석 참조)."""
    path, why = keywordsio.resolve_read(space, _calibration_repo_path(subject))
    if why and report is not None:
        report.note(CALIBRATION_FILE, why, "warn")
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    # 어느 파일에서 읽었는지 리포트가 밝힐 수 있게 실어 보낸다. 보정값은 자동확정
    # 여부를 가르는 값이라 "어느 파일의 0.35 인가"를 되짚을 수 있어야 한다.
    if isinstance(data, dict):
        data.setdefault("_read_from", str(path))
    return data


def _keyword_book(subject, space, report):
    """keywords.json 을 읽는다. 작업공간 사본 우선 — 새 과목이 저장소를 건드리지
    않고도 순환을 돌 수 있어야 하기 때문이다(keywordsio 주석 참조)."""
    path, why = keywordsio.resolve_read(space, subject.keywords_path)
    if why:
        report.note("keywords.json", why, "warn")
    book = keywordsio.load(path, subject.standard_prefixes or {})
    for ident, reason, severity in book.notes:
        report.note(ident, reason, severity)
    return book, path


def _calibration_for(calib: dict, rev: str) -> dict | None:
    """해당 개정의 보정 결과. 자동확정이 허용된 경우에만 돌려준다."""
    row = ((calib.get("revisions") or {}).get(rev)) or {}
    if row.get("auto_confirm") and (row.get("recommended") or {}).get("threshold") is not None:
        # 보정 시각은 파일 최상위에 한 번만 적혀 있다. items 에 "언제 잰 보정이냐"를
        # 남겨야 사전이 바뀐 뒤에 찍힌 분류를 나중에 골라낼 수 있으므로 여기서 합친다.
        return dict(row, calibrated_at=calib.get("calibrated_at"))
    return None


def _confidence_from_curve(row: dict, score: float) -> float | None:
    """원점수 → 실측 정확도의 95% 하한. 곡선에서 '이 점수 이상' 구간을 읽는다.

    모든 자동확정에 같은 숫자를 붙이는 대신 이렇게 하는 이유: 곡선은 임계값별로
    실제로 몇 개가 맞았는지를 이미 재어 두었다. 점수 0.9 짜리와 0.36 짜리에 같은
    confidence 를 붙이면 그 측정을 버리는 셈이다. 반대로 곡선 밖의 숫자를
    지어내지도 않는다 — 표본이 CALIB_MIN_AUTO 미만인 구간은 건너뛴다.
    """
    rec = row.get("recommended") or {}
    floor_th = rec.get("threshold")
    if floor_th is None:
        return None

    # 참조 구간을 고르는 두 가지 규칙, 둘 다 이유가 있다.
    # 1) 권장 임계값 **미만** 구간은 쓰지 않는다. 그 구간에는 애초에 자동확정되지
    #    않았을 문항이 섞여 있어, 지금 자동확정된 이 문항의 동류(reference class)가
    #    아니다. (실측: 지Ⅱ 의 '≥0.20' 구간은 36/36 이지만 그 중 11건은 큐로 갔다.)
    # 2) 자격이 되는 구간이 여럿이면 **가장 센 하한**을 고른다. 임계값을 올리면
    #    표본이 줄어 하한이 도로 내려가기 때문이다(지Ⅱ: 0.35→25건 하한 0.867,
    #    0.45→10건 하한 0.723). 가장 높은 임계값을 고르면 점수가 더 높은 문항에
    #    더 낮은 confidence 가 붙는 뒤집힌 결과가 나온다.
    best = None
    for point in row.get("curve") or []:
        th, lo, n = point.get("threshold"), point.get("accuracy_lo95"), point.get("auto", 0)
        if th is None or lo is None or n < CALIB_MIN_AUTO or th < floor_th:
            continue
        if score >= th and (best is None or lo > best):
            best = float(lo)
    if best is not None:
        return round(best, 4)
    lo = rec.get("accuracy_lo95")
    return round(float(lo), 4) if isinstance(lo, (int, float)) else None


# ---------------------------------------------------------------- classify 본체
def _classify(args) -> int:
    subject = load_subject(args.subject)
    space = Space(subject.slug, getattr(args, "workspace", None)).ensure()
    report = Report("classify", subject.slug, space)

    if not space.items.exists() or not any(space.items.glob("*.json")):
        report.count(total=0, auto=0, queued=0, skipped=0)
        report.note("items", "items/ 가 비어 있다 — 먼저 gw extract 를 실행한다", "warn")
        report.next = f"python scripts/gw.py extract --subject {subject.slug}"
        return report.finish()

    # --- 처리할 개정 목록 결정 ---
    requested = [args.revision] if args.revision != "both" else list(DEFAULT_REVISIONS)
    revisions: list[str] = []
    for rev in requested:
        if not subject.curriculum.get(rev):
            severity = "info" if args.revision == "both" else "error"
            why = (f"curriculum.{rev} 가 정의되지 않아 건너뜀 — 2022는 보통 gw map 으로 "
                   f"2015 분류를 옮겨 만든다" if severity == "info"
                   else f"curriculum.{rev} 가 subject.json 에 정의되지 않았다")
            report.note(rev, why, severity)
            if severity == "error":
                report.next = f"subjects/{subject.slug}/subject.json 의 curriculum 필드를 확인한다"
                return report.finish()
            continue
        revisions.append(rev)
    if not revisions:
        report.note("revision", "처리할 개정이 없다", "error")
        return report.finish()

    only = _selection(getattr(args, "only", None))

    book, book_path = _keyword_book(subject, space, report)
    if book.is_empty():
        report.note("keywords", f"{book_path} 가 없거나 비어 있다 — 전부 큐로 보낸다", "warn")

    # --- items/ 전체를 한 번 읽어 코퍼스 텍스트(IDF용)와 캐시를 동시에 만든다 ---
    items_cache = _load_items(space, report)
    corpus_texts = [_item_text(d) for d in items_cache.values()]

    candidates_by_rev: dict[str, dict[str, list[str]]] = {}
    weights_by_rev: dict[str, dict] = {}
    kw_stats: dict[str, dict] = {}
    for rev in revisions:
        scope = subject.code_scope(rev)
        _note_scope(report, scope)
        cands, weights, stats = _keyword_view(book.revision(rev), scope)
        candidates_by_rev[rev] = cands
        weights_by_rev[rev] = weights
        kw_stats[rev] = stats
        if not cands:
            report.note(rev, f"keywords.json 에 {rev} 개정 층이 없거나 비어 있다 — 전부 큐로 감. "
                             f"python scripts/gw.py standards --draft-keywords "
                             f"--subject {subject.slug} --revision {rev} --force", "warn")

    all_candidate_keywords: set[str] = set()
    for cands in candidates_by_rev.values():
        for kws in cands.values():
            all_candidate_keywords.update(kws)
    idf = _compute_idf(corpus_texts, all_candidate_keywords)

    meta_by_rev = {rev: _load_standards_meta(rev) for rev in revisions}
    if all(not m for m in meta_by_rev.values()):
        report.note("standards", "curriculum/standards/<개정>.json 이 없어 unit 라벨을 못 채운다(code만 기록)", "info")

    # ── 보정 게이트 ──────────────────────────────────────────────────────
    # 보정되지 않은 과목에서는 자동확정을 하지 않는다. 실측된 위험이다:
    # 교육과정 초안 사전으로 자동확정한 판정의 정확도가 40% 였는데 그 오답에
    # 0.94 짜리 점수가 붙어 있었다. 점수를 confidence 로 읽는 순간 items 에
    # 틀린 성취기준이 by=keyword 로 박히고, 아무도 다시 안 본다.
    calib_all = _load_calibration(subject, space, report)
    calib_by_rev: dict[str, dict | None] = {}
    for rev in revisions:
        row = _calibration_for(calib_all, rev)
        calib_by_rev[rev] = row
        if row is None:
            why = (f"{rev}: 이 과목은 아직 보정되지 않았다 — 자동확정을 하지 않고 전부 큐로 보낸다. "
                   f"python scripts/gw.py classify --subject {subject.slug} --calibrate 를 돌리거나, "
                   f"큐 판정 결과를 --apply 한 뒤 --learn 을 돌려 사전을 키워라")
            raw_row = ((calib_all.get("revisions") or {}).get(rev)) or {}
            if raw_row and not raw_row.get("auto_confirm"):
                why = (f"{rev}: 보정 결과가 '자동확정 불가'다 — 목표 정확도 "
                       f"{raw_row.get('target_accuracy', CALIB_TARGET_ACCURACY)} 를 지키는 임계값이 없었다. "
                       f"{raw_row.get('note') or ''}").strip()
            report.note("calibration", why, "warn")
        elif row.get("self_graded"):
            # 자동확정을 켠 그 임계값이 **LLM 라벨만으로** 매겨졌다는 사실은
            # classify 를 돌리는 사람에게도 보여야 한다. calibration.json 안에만
            # 적어 두면 리포트만 읽는 쪽(=이 도구의 유일한 독자, CONTRACT 2절)은
            # "정확도 0.9" 를 사람이 검수한 0.9 로 읽는다.
            report.note("calibration",
                        f"{rev}: 자동확정에 쓰는 임계값 {row['recommended']['threshold']} 은 "
                        f"**LLM 판정 라벨만으로** 매겨졌다(사람 검수 0건). 이 정확도는 "
                        f"'정답과의 일치율'이 아니라 'LLM 판정과의 일치율'이다 — "
                        f"자동확정된 문항을 표본으로 검수해 by=manual 라벨을 늘린 뒤 "
                        f"다시 --calibrate 하면 수치의 뜻이 강해진다", "warn")

    total = auto = queued = skipped = 0
    changed_qids: set[str] = set()
    queue_items: list[dict] = []
    blank_excerpts: set[str] = set()

    # 문항마다 사전 전체와 점수를 맞춰 본다. 380문항 × 개정 2벌이면 체감된다.
    for stem, item in track(items_cache.items(), "문항", total=len(items_cache),
                            label="classify", args=args, detail=lambda kv: kv[0]):
        if not _selected(stem, only):
            continue
        fields = _extract_fields(item)

        for rev in revisions:
            total += 1
            existing = (item.get("classification") or {}).get(rev)
            # 보호 등급: manual(사람 검수) > llm(애매해서 LLM 이 판정) > keyword(이 모듈의
            # 결과). --force 는 "keywords.json 을 고쳤으니 keyword 등급만 다시 계산해줘"
            # 용도로 쓴다 — manual/llm 은 이미 keyword 보다 비싼 검증을 거쳤으므로
            # --force 라도 건드리면 애써 해소한 애매함이 도로 큐에 쌓인다.
            if existing and existing.get("by") in ("manual", "llm"):
                skipped += 1
                continue
            if existing and not args.force:
                skipped += 1
                continue

            candidates = candidates_by_rev[rev]
            scores, matched = _score_item(fields, candidates, idf, weights_by_rev[rev])
            ranked = _rank(scores)
            top1 = ranked[0] if ranked else (None, 0.0)

            calib = calib_by_rev[rev]
            if calib is None:
                # 보정 전에는 임계값을 논할 수 없다. 무조건 큐.
                reasons = ["uncalibrated"]
            else:
                reasons = _queue_reasons(ranked, matched, calib["recommended"]["threshold"])

            if not reasons:
                item.setdefault("classification", {})[rev] = {
                    "standard": top1[0],
                    "unit": _unit_label(meta_by_rev, rev, top1[0]),
                    # confidence 는 '점수'가 아니라 '이 점수대에서 실제로 맞은 비율의
                    # 95% 하한'이다. 보정 곡선에서 읽어온 실측값이고, 정의상 1.0 이
                    # 될 수 없다 — 표본이 작아서 나온 100% 를 확신으로 오독하지 않게.
                    "confidence": _confidence_from_curve(calib, top1[1]),
                    "by": "keyword",
                    # 계약 4절: 계약에 없는 필드는 ext 아래로. 원점수와 보정 시각은
                    # 진단용이지 계약이 보장하는 값이 아니다.
                    "ext": {"score": top1[1], "calibrated_at": calib.get("calibrated_at")},
                }
                auto += 1
                changed_qids.add(stem)
            else:
                # ★ 후보에 성취기준 **문장**을 함께 싣는다.
                #   제작기 화면은 성취기준 전문을 보여주는데 큐를 읽는 LLM 은
                #   코드와 단원명만 받고 있었다 — 판정하는 쪽이 사람보다 적은
                #   정보로 판정하는 상태였다. 문장 없이 `12한지02-02` 만 보고
                #   '대상(명사)으로 판정하라'는 지침을 따를 방법이 없다.
                cand_list = [
                    {"code": c, "unit": _unit_label(meta_by_rev, rev, c),
                     "text": _standard_text(meta_by_rev, rev, c), "score": s}
                    for c, s in ranked[:QUEUE_CANDIDATES] if matched.get(c, 0) > 0
                ]
                excerpt = _build_excerpt(item)
                # 발췌가 비었는데 vision 도 아니면 extract 가 아직 안 돈 것이다.
                # 이런 항목은 LLM 에게 보내봐야 판정할 근거가 없다 — 아래에서 세어 알린다.
                if not excerpt and item.get("extraction_mode") != "vision":
                    blank_excerpts.add(stem)
                queue_items.append({
                    "qid": stem,
                    "revision": rev,
                    "reasons": reasons,
                    "excerpt": excerpt,
                    "candidates": cand_list,
                })
                queued += 1

    if not args.dry_run:
        for stem in sorted(changed_qids):
            space.item(stem).write_text(
                json.dumps(items_cache[stem], ensure_ascii=False, indent=2), encoding="utf-8")

    # 큐 파일은 리포트의 일부다 — dry-run 이어도 (미리보기 목적으로) 쓴다.
    # dry-run 이 막는 것은 items/ 원본 변경뿐이다.
    queue_payload = {
        "step": "classify_queue",
        "slug": subject.slug,
        "guidance": QUEUE_GUIDANCE,
        # ★ 판정 결과 파일의 스키마를 큐가 직접 들고 다닌다.
        #   `classify_result.json` 의 모양이 --help·README·SKILL.md·CONTRACT.md
        #   어디에도 없었다. 그런데 SKILL.md 는 '소스를 먼저 뒤지지 마라'고 못박는다 —
        #   큐를 읽는 LLM 이 계약대로 행동하면 답을 만들 방법이 원리적으로 없었다.
        #   그래서 예시 한 덩이와 필드 설명을 큐 안에 함께 싣는다. 이 파일 하나로
        #   판정 결과를 만들 수 있어야 한다.
        "result_schema": RESULT_SCHEMA,
        # 후보 3개 밖의 코드를 고를 수 있어야 한다. 실측: 지구과학Ⅱ 큐 235건 중
        # 21건은 매칭이 0이라 후보 목록이 **비어 있었다** — 그 문항을 받은 LLM 은
        # 고를 코드 자체를 못 봤다. 문항마다 반복하면 토큰을 먹으니 개정별로
        # 한 번만 싣는다(한국지리 28개 · 지구과학Ⅱ 37개 규모).
        "standards": {
            rev: _standards_catalog(subject, meta_by_rev, rev) for rev in revisions
        },
        "count": len(queue_items),
        "items": queue_items,
    }
    queue_path = space.report("classify_queue")
    queue_path.parent.mkdir(parents=True, exist_ok=True)
    queue_path.write_text(json.dumps(queue_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    report.count(total=total, auto=auto, queued=queued, skipped=skipped)
    report.artifact(space.rel(queue_path))
    report.extra["queue_file"] = space.rel(queue_path)
    # 큐를 통째로 읽을지 나눠 읽을지 정하려면 크기를 알아야 한다. 리포트만 읽는
    # 쪽에게 "열어 보고 판단해라"는 답이 될 수 없다(CONTRACT 2절).
    report.extra["queue_bytes"] = queue_path.stat().st_size
    # 리포트만 읽는 LLM 이 "왜 자동확정이 0인가"를 되묻지 않도록 상태를 실어 보낸다.
    report.extra["calibration"] = {
        rev: ({"threshold": row["recommended"]["threshold"],
               "measured_accuracy": row["recommended"].get("standard_accuracy"),
               # 그 정확도를 무엇으로 쟀는지 없이 숫자만 보이면 사람 검수 결과로 읽힌다.
               "gold_mix": row.get("gold_mix"),
               "self_graded": bool(row.get("self_graded")),
               "calibrated_at": row.get("calibrated_at")} if row else None)
        for rev, row in calib_by_rev.items()
    }
    report.extra["keywords"] = kw_stats
    # 어느 파일을 읽었는지 리포트가 밝힌다 — 작업공간 사본과 저장소 사본이 갈릴 수
    # 있는 이상, "이 판정이 어느 사전에서 나왔나"를 되짚을 수 없으면 안 된다.
    report.extra["keywords_file"] = str(book_path)
    if calib_all.get("_read_from"):
        report.extra["calibration_file"] = calib_all["_read_from"]

    # ── 본문 없는 큐를 조용히 내보내지 않는다 ──────────────────────────────
    # 통합 검증에서 crop 만 돌고 extract 는 안 돈 작업공간에 classify 를 걸었더니
    # 발췌가 전부 빈 399건짜리 큐를 만들고도 ok:true / attention:0 으로 끝났다.
    # 그 리포트만 읽은 LLM 은 판정 근거가 한 글자도 없는 큐를 그대로 물게 된다.
    # vision 문항은 원래 본문이 없으니(크롭 이미지가 본체) 여기서 제외한다.
    if blank_excerpts:
        exams = sorted({split_qid(q)[0] for q in blank_excerpts})
        shown = ", ".join(exams[:8]) + (" 외" if len(exams) > 8 else "")
        report.note("extract", f"큐에 담긴 문항 중 {len(blank_excerpts)}건은 본문이 비어 있다 "
                               f"(vision 아님) — extract 가 아직 안 돌았을 수 있다. 회차: {shown}",
                    "warn")

    if blank_excerpts:
        # 근거 없는 큐를 LLM 에게 넘기기 전에 extract 로 되돌린다.
        report.next = f"python scripts/gw.py extract --subject {subject.slug}"
    elif queued:
        # 사전은 쓸수록 좋아진다 — 큐 판정을 --apply 한 뒤 --learn 을 돌리면
        # 다음 회차부터 큐가 줄어든다. 그 순환을 next 에 명시한다.
        # 보정 안 된 과목은 auto=0 으로 전부 큐에 담긴다(정상 동작). 그 사람이 원한 게
        # 학습지 한 장이면 여기서 판정을 기다릴 이유가 없다는 것도 함께 적는다 —
        # build 는 미분류 문항을 회차별 '미분류' 탭으로 내보낸다.
        report.next = (
            f"LLM 이 {space.rel(queue_path)} 을 읽고 classify_result.json 을 만든 뒤 "
            f"python scripts/gw.py classify --subject {subject.slug} --apply classify_result.json "
            f"→ 이어서 --learn 과 --calibrate 를 돌리면 사전과 임계값이 이 과목에 맞게 갱신된다. "
            f"단원 분류 없이 학습지만 필요하면 "
            f"python scripts/gw.py build --subject {subject.slug} 로 바로 가도 된다"
        )
    else:
        report.next = f"python scripts/gw.py map --subject {subject.slug}"

    if args.quiet:
        report.write()
        return 1 if report.has_error else 0
    return report.finish()


# ---------------------------------------------------------------- --learn
def _learn(args) -> int:
    """라벨된 문항에서 사전을 배운다. 기존 keywords.json 과 병합해 다시 쓴다."""
    subject = load_subject(args.subject)
    space = Space(subject.slug, getattr(args, "workspace", None)).ensure()
    report = Report("classify_learn", subject.slug, space)

    revisions = ([args.revision] if args.revision != "both"
                 else [r for r in DEFAULT_REVISIONS if subject.curriculum.get(r)])
    if not revisions:
        report.note("revision", "학습할 개정이 없다 — subject.json 의 curriculum 을 확인해라", "error")
        return report.finish()

    items = _load_items(space, report)
    if not items:
        report.note("items", "items/ 가 비어 있다 — 먼저 gw extract 를 실행한다", "error")
        report.next = f"python scripts/gw.py extract --subject {subject.slug}"
        return report.finish()

    only = _selection(getattr(args, "only", None))
    book, book_path = _keyword_book(subject, space, report)
    learned_from_all: dict[str, list[str]] = dict((book.meta.get("learned_from") or {}))

    n_terms = 0
    diag_by_rev: dict[str, dict] = {}
    mix_by_rev: dict[str, dict] = {}
    for rev in revisions:
        scope = subject.code_scope(rev)
        _note_scope(report, scope)
        # 학습 라벨은 manual(사람) 과 llm(큐 판정) 을 모두 받는다.
        # 채점(--calibrate)도 이제 둘 다 받되, **표본의 구성을 밝히고** 사람 라벨이
        # 0건이면 '자기 채점'이라고 경고한다 — 아래 `_calibrate` 주석 참조.
        # 라벨 코드를 scope 로 거르는 이유: 다른 개정의 코드가 라벨에 섞여 있으면
        # 그 용어가 이 개정 사전에 학습돼 다음 실행부터 **에러 없이** 남의 개정
        # 성취기준을 추천하게 된다. 접두사 비교로는 통합과목처럼 접두사가 겹치는
        # 자리에서 못 걸러냈다.
        gold, label_source = _gold_labels(subject, items, rev, ("manual", "llm"))
        out_of_scope = sorted({q for q, c in gold.items()
                               if _selected(q, only) and c not in scope})
        gold = {q: c for q, c in gold.items()
                if _selected(q, only) and c in scope}
        if out_of_scope:
            # 조용히 버리면 "라벨 40건 넣었는데 왜 20건만 배웠나"를 못 되짚는다.
            report.note(rev, f"{rev}: 라벨 {len(out_of_scope)}건은 이 개정 성취기준이 아니라 "
                             f"제외했다({', '.join(out_of_scope[:4])}"
                             f"{'…' if len(out_of_scope) > 4 else ''})", "warn")
        # 출처 집계는 --only 로 걸러낸 **뒤** 다시 센다. 필터 전 숫자를 리포트에 실으면
        # "라벨 120건(items(by=manual)=160)" 처럼 앞뒤가 안 맞는 문장이 나온다.
        sources = _label_mix(label_source, gold)
        mix_by_rev[rev] = sources
        if not gold:
            report.note(rev, f"{rev}: 라벨된 문항이 없다 — 큐 판정을 --apply 하거나 "
                             f"gw map 으로 mapping.json 을 주입한 뒤 다시 돌려라", "warn")
            continue
        if len(gold) < LEARN_MIN_ITEMS:
            report.note(rev, f"{rev}: 라벨 {len(gold)}건뿐이다(권장 최소 {LEARN_MIN_ITEMS}건). "
                             f"배운 용어가 우연일 수 있다 — 큐를 더 판정한 뒤 다시 돌려라", "warn")

        # 문항마다 본문을 통째로 토큰화한다 — 라벨이 수백 건이면 눈에 띄게 걸린다.
        docs = {q: _terms_of(_item_text(items[q]))
                for q in track(list(gold), "문항", label="classify --learn",
                               args=args, detail=str)}
        docs = {q: t for q, t in docs.items() if t}   # 본문 없는 vision 문항은 배울 게 없다
        skipped_blank = len(gold) - len(docs)
        if skipped_blank:
            report.note(rev, f"{rev}: 본문이 비어 학습에서 제외한 문항 {skipped_blank}건 "
                             f"(vision 모드는 크롭 이미지가 본체라 정상이다)", "info")

        learned, diag = _mine_terms(docs, gold)
        diag["labels"] = len(gold)
        diag["blank_text"] = skipped_blank
        diag_by_rev[rev] = diag
        if not learned:
            # ★ 조용한 무동작 금지. 라벨이 있는데 배운 게 0이면 **왜 0인지** 말한다.
            #   실측: 한국지리 라벨 20건에서 learned_terms=0 이 나왔고, 리포트에는
            #   "배운 것이 없다"만 있어서 다음에 뭘 해야 할지 알 수 없었다.
            report.note(rev, f"{rev}: 라벨 {len(gold)}건을 봤지만 배운 용어가 0개다 — "
                             + _learn_why_zero(diag), "warn")

        # --- 병합: 교육과정 유래는 보존하고, learned 만 통째로 갈아끼운다 ---
        # 덧붙이지 않고 교체하는 이유: --learn 은 언제든 다시 돌 수 있어야 하고,
        # 돌 때마다 옛 용어가 쌓이면 라벨이 늘어도 사전이 낡은 채로 굳는다.
        # ★ **이 개정 층만** 건드린다. 다른 개정의 사전은 이 실행의 라벨과 아무 관계가
        #   없으므로 손대면 안 된다(같은 코드를 두 개정이 쓰는 과목에서 특히 중요하다).
        for code, entry in list(book.revision(rev).items()):
            if code not in scope:
                continue
            book.set_entry(rev, code, {"curriculum": list(entry.get("curriculum") or []),
                                       "learned": learned.get(code, [])})
        for code, rows in learned.items():
            if code not in book.revision(rev):
                book.set_entry(rev, code, {"curriculum": [], "learned": rows})

        # ── 사전이 몇 개 코드에만 붙으면 그쪽으로 쏠린다 ──────────────────
        # 실측(한국지리, 라벨 20건): 18개 성취기준 중 문항이 2건 이상 모인 코드가
        # 2개뿐이라 용어 17개가 그 2개에만 붙었고, 홀드아웃 20건 성취기준 정확도가
        # 0.10 → 0.05 로 **내려갔다**(top-3 은 0.10 → 0.20 으로 올랐다).
        # 학습이 항상 이득이라고 말하면 안 된다 — 라벨이 얇으면 손해일 수 있고,
        # 그 판정은 --calibrate 가 홀드아웃에서 재는 것이 유일한 근거다.
        if learned and len(learned) * 3 < diag["codes"]:
            report.note(rev, f"{rev}: 배운 용어가 성취기준 {len(learned)}개에만 붙었다"
                             f"(라벨이 걸린 성취기준은 {diag['codes']}개). 그 몇 개 코드로 "
                             f"판정이 쏠릴 수 있다 — --calibrate 를 홀드아웃 회차로 돌려 "
                             f"정확도가 실제로 올랐는지 확인해라. 내려가면 라벨을 더 모은 뒤 "
                             f"다시 --learn 하는 편이 낫다", "warn")

        learned_from_all[rev] = sorted(docs)
        n_terms += sum(len(v) for v in learned.values())
        report.count(**{f"labeled_{rev}": len(gold), f"codes_{rev}": len(learned),
                        f"terms_{rev}": sum(len(v) for v in learned.values())})
        report.note(rev, f"{rev}: 라벨 {len(gold)}건({', '.join(f'{k}={v}' for k, v in sources.items())})에서 "
                         f"{len(learned)}개 성취기준의 용어 {sum(len(v) for v in learned.values())}개를 배웠다. "
                         f"라벨이 없는 성취기준은 교육과정 초안 키워드만 남는다", "info")

    book.meta.update({
        # schema 문자열은 사람이 읽는 메모일 뿐이다. 층을 가르는 판정은 **모양**으로
        # 한다(keywordsio 참조) — 스키마 이름을 믿었다가 조용히 틀린 전례가 있다.
        "schema": keywordsio.SCHEMA,
        "learned_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "learned_by": "gw classify --learn (로그 오즈비)",
        "params": {"min_df": LEARN_MIN_DF, "max_df_ratio": LEARN_MAX_DF_RATIO,
                    "min_lor": LEARN_MIN_LOR, "top_per_code": LEARN_TOP_PER_CODE,
                    "alpha": LEARN_ALPHA, "shrink_k": LEARN_SHRINK_K},
        # 어느 문항으로 배웠는지 남긴다. --calibrate 가 이 목록을 빼고 채점해야
        # "학습에 쓴 문항으로 자기를 채점하는" 사고를 막을 수 있다.
        "learned_from": learned_from_all,
        # 이 사전이 어떤 문항으로 만들어졌는지도 파일이 스스로 들고 있어야 한다.
        "learned_from_exams": {
            rev: sorted({split_qid(q)[0] for q in qids if _is_qid(q)})
            for rev, qids in learned_from_all.items()
        },
    })

    # 깔때기 진단은 성공했을 때도 싣는다 — "왜 이만큼만 배웠나"도 같은 숫자로 읽힌다.
    report.extra["learn_funnel"] = diag_by_rev
    report.extra["label_sources"] = mix_by_rev

    if not learned_from_all:
        # 아무것도 못 배웠는데 파일을 만들면 `gw subjects` 의 readiness 가 "keywords 있음"
        # 으로 켜진다 — 내용이 _meta 뿐인 껍데기를 사전이 있다고 보고하는 셈이다.
        report.note("keywords", "배운 것이 없어 keywords.json 을 쓰지 않았다", "warn")
        report.count(learned_terms=0)
        report.next = (f"큐를 판정해 --apply 하거나 gw map 으로 라벨을 주입한 뒤 "
                       f"python scripts/gw.py classify --subject {subject.slug} --learn 를 다시 돌려라")
        return report.finish()

    # 저장소가 아니라 작업공간에 쓴다(keywordsio 주석: git pull 이 막히는 자리).
    out_path = keywordsio.write_target(space, subject.keywords_path,
                                       bool(getattr(args, "install", False)))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    keywordsio.save(out_path, book, dry_run=bool(args.dry_run))

    report.artifact(str(out_path))
    report.count(learned_terms=n_terms)
    if not getattr(args, "install", False):
        report.note("keywords.json",
                    keywordsio.install_hint(space, subject.keywords_path,
                                            f"classify --subject {subject.slug} --learn"),
                    "info")
    # 사전이 바뀌면 옛 보정값은 더 이상 그 사전의 것이 아니다.
    report.note("calibration", "사전이 바뀌었으므로 임계값을 다시 재야 한다 — "
                                "--calibrate 를 돌리기 전까지 자동확정은 옛 보정값을 쓴다", "warn")
    report.next = (f"python scripts/gw.py classify --subject {subject.slug} --calibrate "
                   f"--only <학습에 쓰지 않은 회차>")
    return report.finish()


# ---------------------------------------------------------------- --calibrate
def _calibrate(args) -> int:
    """라벨(manual·llm)로 자기 정확도를 재고 권장 임계값을 calibration.json 에 쓴다.

    실패하면(잴 라벨이 없으면) 파일을 **쓰지 않는다** — 아래 마지막 블록 주석 참조.
    """
    subject = load_subject(args.subject)
    space = Space(subject.slug, getattr(args, "workspace", None)).ensure()
    report = Report("classify_calibrate", subject.slug, space)

    revisions = ([args.revision] if args.revision != "both"
                 else [r for r in DEFAULT_REVISIONS if subject.curriculum.get(r)])
    if not revisions:
        report.note("revision", "보정할 개정이 없다 — subject.json 의 curriculum 을 확인해라", "error")
        return report.finish()

    items = _load_items(space, report)
    if not items:
        report.note("items", "items/ 가 비어 있다 — 먼저 gw extract 를 실행한다", "error")
        report.next = f"python scripts/gw.py extract --subject {subject.slug}"
        return report.finish()

    only = _selection(getattr(args, "only", None))
    book, book_path = _keyword_book(subject, space, report)
    trained_on = book.meta.get("learned_from") or {}
    corpus_texts = [_item_text(d) for d in items.values()]

    payload = {
        "step": "calibrate",
        "slug": subject.slug,
        "calibrated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "target_accuracy": CALIB_TARGET_ACCURACY,
        "contract_floor": TOP_MIN,
        "gap_min": GAP_MIN,
        # 어떤 자료로 잰 숫자인지 파일 자체가 들고 있어야 한다. 이 파일은 저장소에
        # 커밋되어 다른 사람의 작업공간에서 읽히므로, 출처가 없으면 "이 0.35 는
        # 어디서 나온 값인가"를 아무도 되짚을 수 없다.
        "measured_on": {"items": len(items),
                         "exams": sorted({split_qid(q)[0] for q in items if _is_qid(q)})},
        "revisions": {},
    }

    for rev in revisions:
        scope = subject.code_scope(rev)
        _note_scope(report, scope)
        candidates, weights, stats = _keyword_view(book.revision(rev), scope)
        meta_by_rev = {rev: _load_standards_meta(rev)}
        unit_of = {code: _unit_label(meta_by_rev, rev, code) for code in candidates}

        # ── 채점 정답으로 무엇을 받을 것인가 ──────────────────────────────
        # 예전에는 `manual` 만 받았다. 그 규칙은 원리적으로 순환을 끊는다:
        # `--apply` 는 라벨을 `by=llm` 으로 쓰고, `by=manual` 을 만드는 코드는
        # `gw map` 이 mapping.json 을 읽을 때뿐인데 mapping.json 은 4과목에만
        # 있다. 즉 새 과목은 무엇을 해도 보정에 도달할 수 없었다 —
        # README·SKILL.md 가 제시한 --apply → --learn → --calibrate 순환이
        # 문서에만 있고 코드에는 없던 자리다.
        #
        # 그래서 llm 라벨도 받는다. 대신 **측정의 뜻을 흐리지 않는다.**
        #   ① 학습에 쓴 문항은 계속 뺀다(`learned_from` 홀드아웃). 같은 문항으로
        #      배우고 같은 문항으로 채점하는 것이 진짜 '자기 채점'이고, 그 구멍은
        #      라벨 출처와 무관하게 막혀 있어야 한다.
        #   ② 표본의 구성(manual/llm 몇 건)을 calibration.json 과 리포트에 그대로
        #      싣는다. 숫자 옆에 출처가 없으면 읽는 쪽은 사람 검수 결과로 읽는다.
        #   ③ 사람 라벨이 0건이면 `self_graded: true` 를 남기고 경고한다. 그 정확도는
        #      '정답과의 일치율'이 아니라 'LLM 판정과의 일치율'이다 — 사전이 그 LLM
        #      판정에서 학습됐다면, 같은 판단 습관을 되풀이하는 오류는 이 채점으로
        #      절대 드러나지 않는다.
        #   ④ 그래도 자동확정 자체는 연다. 여기서 막으면 새 과목은 여전히 막다른
        #      골목이고, '측정했지만 못 믿는다'와 '측정조차 못 한다' 중 전자가 낫다.
        #      다만 `classify` 가 자동확정할 때마다 ③의 경고를 다시 띄운다.
        gold, label_source = _gold_labels(subject, items, rev, ("manual", "llm"))
        gold = {q: c for q, c in gold.items() if _selected(q, only)}
        sources = _label_mix(label_source, gold)
        if not gold:
            report.note(rev, f"{rev}: 채점할 라벨이 하나도 없다 — 큐를 판정해 --apply 하거나 "
                             f"gw map 으로 mapping.json 을 주입한 뒤 다시 돌려라", "error")
            continue

        used_in_training = set(trained_on.get(rev) or [])
        holdout = {q: c for q, c in gold.items() if q not in used_in_training}
        leaked = len(gold) - len(holdout)
        scored_on = holdout if holdout else gold
        n_human, n_llm = _human_share(label_source, scored_on)
        self_graded = n_human == 0

        idf = _compute_idf(corpus_texts, {t for v in candidates.values() for t in v})

        def measure(subset: dict[str, str]) -> tuple[dict, list[dict]]:
            rows = []
            # 채점은 문항마다 사전 전체를 다시 훑는다. 임계값 곡선까지 돌면 오래 걸린다.
            for qid, truth in track(subset.items(), "문항", total=len(subset),
                                    label="classify --calibrate", args=args,
                                    detail=lambda kv: kv[0]):
                scores, matched = _score_item(_extract_fields(items[qid]), candidates, idf, weights)
                rows.append((truth, _rank(scores), matched))
            n = len(rows)
            if not n:
                return {"n": 0}, []
            std_ok = sum(1 for t, r, _ in rows if r and r[0][0] == t)
            # 단원 정확도를 따로 재는 이유: 실측상 성취기준은 틀려도 단원은 맞는
            # 경우가 대부분이다(오답의 절반 이상). 단원까지 못 맞히는 것과
            # 단원 안에서 형제를 헷갈리는 것은 대응이 다르다.
            unit_ok = sum(1 for t, r, _ in rows
                          if r and unit_of.get(r[0][0]) and unit_of.get(r[0][0]) == unit_of.get(t))
            top3_ok = sum(1 for t, r, _ in rows if t in [c for c, _s in r[:QUEUE_CANDIDATES]])
            overall = {"n": n,
                        "standard_accuracy": round(std_ok / n, 4),
                        "standard_accuracy_lo95": _wilson_lower(std_ok, n),
                        "unit_accuracy": round(unit_ok / n, 4),
                        "top3_accuracy": round(top3_ok / n, 4)}
            curve = []
            for th in CALIB_THRESHOLDS:
                a = a_std = a_unit = q_n = q_top3 = 0
                for truth, ranked, matched in rows:
                    if _queue_reasons(ranked, matched, th):
                        q_n += 1
                        if truth in [c for c, _s in ranked[:QUEUE_CANDIDATES]]:
                            q_top3 += 1
                        continue
                    a += 1
                    if ranked[0][0] == truth:
                        a_std += 1
                    if unit_of.get(ranked[0][0]) == unit_of.get(truth):
                        a_unit += 1
                curve.append({
                    "threshold": th, "auto": a, "auto_rate": round(a / n, 4),
                    "standard_accuracy": round(a_std / a, 4) if a else None,
                    # 점추정만 보면 10건 중 10건에서 1.0 이 나온다. 하한을 같이 실어야
                    # 그 1.0 이 표본 크기의 산물임을 읽는 쪽이 알 수 있다.
                    "accuracy_lo95": _wilson_lower(a_std, a) if a else None,
                    "unit_accuracy": round(a_unit / a, 4) if a else None,
                    "queued": q_n,
                    "queued_top3_accuracy": round(q_top3 / q_n, 4) if q_n else None,
                })
            return overall, curve

        overall, curve = measure(scored_on)
        all_overall, _all_curve = measure(gold)

        # --- 권장 임계값: 목표 정확도를 지키는 가장 낮은 값 ---
        # CONTRACT 7절이 못박은 0.35 아래로는 내려가지 않는다. 계약보다 느슨해지는
        # 방향으로 보정이 작동하면 안 된다(보정은 자동확정을 더 조이기 위한 장치다).
        recommended = None
        for point in curve:
            if point["threshold"] < TOP_MIN:
                continue
            if point["auto"] < CALIB_MIN_AUTO:
                continue
            if point["standard_accuracy"] is not None and point["standard_accuracy"] >= CALIB_TARGET_ACCURACY:
                recommended = point
                break

        note = ""
        auto_ok = recommended is not None
        if not auto_ok:
            note = (f"목표 정확도 {CALIB_TARGET_ACCURACY} 를 자동확정 표본 {CALIB_MIN_AUTO}건 이상으로 "
                    f"지키는 임계값이 없다 — 이 과목·이 사전으로는 자동확정 불가. 전부 큐로 보낸다. "
                    f"큐 판정을 --apply 한 뒤 --learn 을 돌려 사전을 키우고 다시 보정해라.")
        # 학습에 쓴 문항으로 채점하면 정확도가 부풀려진다. 부풀려진 숫자로 자동확정을
        # 열어주면 보정을 안 한 것보다 나쁘다 — 틀린 확신을 주기 때문이다.
        # 표본 하한은 학습 여부와 무관하게 건다. 15건짜리 채점으로 임계값을 권장하는
        # 것은, 학습을 안 했더라도 여전히 근거가 모자란 일이다.
        if not holdout:
            auto_ok = False
            recommended = None
            note = ("라벨 전부가 학습에 쓰였다 — 채점할 표본이 남지 않았다. "
                    "--learn 을 --only 로 회차 일부만 주고 다시 돌려라.")
        elif len(holdout) < CALIB_MIN_HOLDOUT:
            auto_ok = False
            recommended = None
            note = (f"채점 표본이 {len(holdout)}건뿐이다(최소 {CALIB_MIN_HOLDOUT}건) — 자동확정을 막는다. "
                    + (f"학습에 쓴 {leaked}건은 채점에서 뺐다. --learn 을 --only 로 회차 일부만 주고 "
                       f"다시 돌려 채점용 회차를 남겨라." if leaked
                       else "큐를 더 판정해 라벨을 늘린 뒤 다시 보정해라."))

        payload["revisions"][rev] = {
            "gold_count": len(gold),
            "gold_sources": sources,
            # ★ 이 숫자가 무엇과의 일치율인지 파일이 스스로 밝힌다.
            #   gold_mix 는 **채점에 실제로 쓴 표본**(holdout)의 구성이다 —
            #   gold_sources 는 --only 로 자르기 전 전체라 둘이 다를 수 있다.
            "gold_mix": {"manual": n_human, "llm": n_llm},
            "self_graded": bool(self_graded),
            "measured_against": ("llm 판정(사람 검수 0건)" if self_graded
                                 else ("사람 검수 라벨" if not n_llm else "사람 검수 + llm 판정 혼합")),
            # overall 이 어떤 표본에서 나온 숫자인지 파일이 스스로 밝힌다.
            "evaluated_on": "holdout" if holdout else "all_gold(학습 데이터 포함 — 신뢰 불가)",
            "holdout_count": len(holdout),
            "used_in_training": leaked,
            "keywords": stats,
            "overall": overall,
            "overall_all_gold": all_overall,
            "curve": curve,
            "recommended": recommended,
            "auto_confirm": bool(auto_ok),
            "note": note,
        }

        sev = "info" if auto_ok else "warn"
        head = (f"{rev}: 성취기준 정확도 {overall.get('standard_accuracy')} / "
                f"단원 정확도 {overall.get('unit_accuracy')} / top-3 {overall.get('top3_accuracy')} "
                f"(채점 {overall.get('n')}건")
        head += f", 학습에 쓰인 {leaked}건 제외)" if leaked else ")"
        # 정확도 바로 옆에 '무엇과의 일치율인가'를 붙인다. 떨어뜨려 놓으면 안 읽힌다.
        head += f" · 채점 기준 라벨 manual {n_human}건 / llm {n_llm}건"
        if auto_ok:
            head += (f" · 권장 임계값 {recommended['threshold']} → 자동확정률 "
                     f"{recommended['auto_rate']}({recommended['auto']}건), 그 구간 실측 정확도 "
                     f"{recommended['standard_accuracy']} (95% 하한 {recommended['accuracy_lo95']} "
                     f"— items 의 confidence 로 이 하한이 실린다)")
        else:
            head += " · 자동확정 불가"
        report.note(rev, head + (f" — {note}" if note else ""), sev)
        if self_graded:
            report.note(rev, f"{rev}: ⚠ 자기 채점이다 — 채점 기준이 전부 LLM 판정 라벨이고 "
                             f"사전도 그 라벨에서 배웠다. 여기 적힌 정확도는 '정답과의 일치율'이 "
                             f"아니라 'LLM 판정과의 일치율'이며, 같은 판단 습관에서 나온 오류는 "
                             f"이 채점으로 드러나지 않는다. 자동확정된 문항을 표본으로 검수해 "
                             f"by=manual 로 남긴 뒤 다시 --calibrate 하면 수치의 뜻이 강해진다",
                        "warn")
        report.count(**{f"gold_{rev}": len(gold), f"holdout_{rev}": len(holdout)})

    # ── 실패한 실행은 저장소도 작업공간도 더럽히지 않는다 ────────────────
    # 실측 사고: 라벨이 하나도 없는 과목에서 [FAIL] 로 끝난 --calibrate 가
    # `revisions: {}` 짜리 calibration.json 을 남겼다. 그 껍데기는 (a) 저장소에
    # 추적되지 않는 파일로 남아 git pull 을 막고, (b) 다음 실행에서 '보정된 과목'
    # 처럼 보이는 파일이 되고, (c) `gw subjects` 의 준비도 표를 흐린다.
    # 아무것도 못 잰 보정 결과는 결과가 아니다.
    ok = (not report.has_error) and bool(payload["revisions"])
    path = keywordsio.write_target(space, _calibration_repo_path(subject),
                                   bool(getattr(args, "install", False)))
    if ok and not args.dry_run:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        report.artifact(str(path))
        report.extra["calibration_file"] = str(path)
    else:
        report.note("calibration.json",
                    f"보정에 실패해서 {path} 를 쓰지 않았다 — 빈 보정 파일이 남으면 "
                    f"다음 실행이 '보정된 과목'으로 착각한다", "warn")

    if ok and not getattr(args, "install", False):
        report.note("calibration.json",
                    keywordsio.install_hint(space, _calibration_repo_path(subject),
                                            f"classify --subject {subject.slug} --calibrate"),
                    "info")

    any_auto = any(r.get("auto_confirm") for r in payload["revisions"].values())
    report.next = (f"python scripts/gw.py classify --subject {subject.slug}" if any_auto else
                   f"python scripts/gw.py classify --subject {subject.slug}  "
                   f"# 자동확정 없이 전부 큐로 간다. 큐를 판정·--apply 한 뒤 --learn → --calibrate 를 반복해라")
    return report.finish()


# ---------------------------------------------------------------- --apply
def _apply(args) -> int:
    subject = load_subject(args.subject)
    space = Space(subject.slug, getattr(args, "workspace", None)).ensure()
    report = Report("classify_apply", subject.slug, space)

    result_path = Path(args.apply)
    if not result_path.exists():
        report.note("apply", f"결과 파일이 없다: {result_path}", "error")
        return report.finish()
    try:
        payload = json.loads(result_path.read_text(encoding="utf-8"))
    except Exception as exc:
        report.note("apply", f"json 파싱 실패: {exc}", "error")
        return report.finish()

    if isinstance(payload, dict) and "results" in payload:
        results = payload["results"]
    elif isinstance(payload, list):
        results = payload
    else:
        report.note("apply", "결과 파일은 배열이거나 {'results': [...]} 형태여야 한다", "error")
        return report.finish()

    applied = errors = 0
    touched: set[tuple[str, str]] = set()
    # 성취기준 메타·범위는 개정마다 한 번만 읽는다(결과 파일이 400줄이어도 파일은 두 번).
    meta_cache: dict[str, dict] = {}
    scope_cache: dict[str, object] = {}
    derived_units = 0

    for row in results:
        qid = row.get("qid") if isinstance(row, dict) else None
        rev = row.get("revision") if isinstance(row, dict) else None
        standard = row.get("standard") if isinstance(row, dict) else None
        if not qid or not rev or not standard:
            report.note(str(qid or "?"), "qid/revision/standard 중 누락됨", "error")
            errors += 1
            continue

        rev = str(rev)
        # 이 과목에 없는 개정으로 들어온 판정은 받지 않는다. korean-geography 는
        # 2022 에 대응 과목이 없어(curriculum.2022 = null) 2022 라벨이 들어오면
        # 갈 곳이 없는데, 예전에는 그대로 items 에 박혔다.
        if not subject.curriculum.get(rev):
            report.note(qid, f"revision {rev} 는 이 과목에 정의돼 있지 않다 "
                             f"(subject.json 의 curriculum.{rev} 가 비어 있다)", "error")
            errors += 1
            continue
        if rev not in meta_cache:
            meta_cache[rev] = _load_standards_meta(rev)
            scope_cache[rev] = subject.code_scope(rev)

        # ── 없는 성취기준 코드를 받아들이지 않는다 ────────────────────────
        # 실측: `12한지99-99` 같은 실재하지 않는 코드가 errors=0 으로 통과해
        # items 에 by=llm 으로 박혔다. 그 문항은 제작기에서 '(단원 미분류)' 로
        # 빠지는데, 리포트는 성공이라 아무도 되짚지 않는다. '틀린 값을 채우느니
        # 비워 둔다'(CONTRACT 0절 5항)의 정확한 반대 방향이라 막는다.
        # 범위 판정은 subject.code_scope 한 곳에서만 한다 — 접두사 문자열 비교는
        # 접두사가 겹치는 과목에서 남의 개정 코드를 통과시킨다(CONTRACT 3절).
        scope = scope_cache[rev]
        known = meta_cache[rev]
        if known and standard not in known:
            report.note(qid, f"성취기준 코드 {standard} 가 curriculum/standards/{rev}.json 에 "
                             f"없다 — 오타이거나 개정이 다르다. 이 문항은 건드리지 않았다", "error")
            errors += 1
            continue
        if standard not in scope:
            report.note(qid, f"성취기준 코드 {standard} 는 이 과목의 {rev} 개정 범위 밖이다 "
                             f"(다른 과목·다른 개정 코드). 이 문항은 건드리지 않았다", "error")
            errors += 1
            continue

        item_path = space.item(qid)
        if not item_path.exists():
            report.note(qid, "items/ 에 해당 문항이 없다", "error")
            errors += 1
            continue
        try:
            item = json.loads(item_path.read_text(encoding="utf-8"))
        except Exception as exc:
            report.note(qid, f"items json 파싱 실패: {exc}", "error")
            errors += 1
            continue

        # ── unit 을 생략하면 코드에서 유도한다 ────────────────────────────
        # 실측: unit 없이 --apply 하면 제작기 사이드바가 통째로 '(단원 미분류)' 가
        # 되고 그 문자열이 PDF 파일명에까지 들어간다. 코드만 알면 단원은
        # curriculum/standards 에서 결정되는 값이라, 사람에게 다시 물을 이유가 없다.
        # 라벨 형식(`(3) 한반도의 지질`)도 여기서 맞춰야 build 가 같은 단원을 두
        # 칸으로 가르지 않는다(_unit_text 주석 참조).
        unit = row.get("unit")
        if not unit:
            unit = _unit_label(meta_cache, rev, standard)
            if unit:
                derived_units += 1

        entry = {
            "standard": standard,
            "unit": unit,
            # ★ confidence 는 '보정으로 실측한 정확도의 95% 하한' 자리다(CONTRACT 4절).
            #   LLM 이 스스로 매긴 숫자를 그 칸에 넣으면, 이 모듈이 고발한 "틀린 답에
            #   붙은 0.96" 을 판정 단계에서 되풀이하게 된다. 받되 ext 로 내린다.
            "confidence": None,
            "by": "llm",
        }
        if isinstance(row.get("confidence"), (int, float)) and not isinstance(row.get("confidence"), bool):
            entry["ext"] = {"llm_confidence": row["confidence"]}
        item.setdefault("classification", {})[rev] = entry
        if row.get("notes"):
            item.setdefault("notes", []).append(f"classify(llm): {row['notes']}")

        if not args.dry_run:
            item_path.write_text(json.dumps(item, ensure_ascii=False, indent=2), encoding="utf-8")
        applied += 1
        touched.add((qid, rev))

    # 큐에서 처리 완료분을 뺀다 — LLM 이 배치로 나눠 여러 번 --apply 하는
    # 반복 워크플로를 가정한다(CONTRACT 7절: "LLM 은 판정 결과를 되돌려주고
    # --apply 가 그것을 items/ 에 반영한다").
    queue_path = space.report("classify_queue")
    if queue_path.exists() and not args.dry_run and touched:
        try:
            queue = json.loads(queue_path.read_text(encoding="utf-8"))
            remaining = [q for q in queue.get("items", [])
                         if (q.get("qid"), q.get("revision")) not in touched]
            queue["items"] = remaining
            queue["count"] = len(remaining)
            queue_path.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            report.note("queue", f"classify_queue.json 갱신 실패(수동 확인 필요): {exc}", "warn")

    report.count(applied=applied, errors=errors, units_derived=derived_units)
    if errors:
        # 거부된 행이 있으면 그 문항은 큐에 그대로 남아 있다. 다음 행동은
        # '고쳐서 다시 --apply' 이지 map 이 아니다.
        report.next = (f"거부된 {errors}건의 성취기준 코드를 고쳐 다시 "
                       f"python scripts/gw.py classify --subject {subject.slug} "
                       f"--apply {result_path}")
    else:
        report.next = (f"python scripts/gw.py classify --subject {subject.slug} --learn "
                       f"→ --calibrate 로 사전과 임계값을 이 과목에 맞게 갱신한다"
                       if applied else None)
    return report.finish(ok=(errors == 0))


# ---------------------------------------------------------------- 엑셀 왕복
def _export_xlsx(args) -> int:
    subject = load_subject(args.subject)
    space = Space(subject.slug, getattr(args, "workspace", None)).ensure()
    report = Report("classify_export_xlsx", subject.slug, space)

    try:
        from openpyxl import Workbook
    except ImportError:
        report.note("openpyxl", "openpyxl 이 설치되어 있지 않다 (requirements.txt 에는 있다 — pip install 확인)", "error")
        return report.finish()

    revisions = list(DEFAULT_REVISIONS) if args.revision == "both" else [args.revision]
    book, _book_path = _keyword_book(subject, space, report)

    # ── 개정 × 코드가 열이 된다 ────────────────────────────────────────────
    # 예전에는 코드만이 열이었다. 그러면 두 개정이 같은 코드를 쓰는 과목(윤리와 사상
    # 등 5개)에서 한 열이 두 교육과정을 가리키게 되고, 되돌릴 때 어느 개정에 써야
    # 할지 알 수 없다. 그래서 1행에 개정을 싣는다.
    columns: list[tuple[str, str]] = []      # (개정, 코드)
    meta_by_rev: dict[str, dict] = {}
    for rev in revisions:
        meta_by_rev[rev] = _load_standards_meta(rev)
        scope = subject.code_scope(rev)
        _note_scope(report, scope)
        # `_load_standards_meta(rev)` 는 그 개정 파일 **전체**(2022 는 2,537개)를 준다.
        # 접두사로 걸렀을 때는 통합과목처럼 접두사가 겹치는 자리에서 남의 개정 코드가
        # 열로 섞여 들어왔다. 사람이 채우는 시트에 그런 열이 있으면 거기 적힌 용어가
        # --import-xlsx 로 그대로 이 개정 사전에 들어앉는다.
        codes = scope.filter(set(book.revision(rev)) | set(meta_by_rev[rev]))
        columns.extend((rev, c) for c in sorted(codes))

    if not columns:
        report.note("codes", "내보낼 성취기준 코드가 없다(keywords.json 도, curriculum/standards 도 비어 있음)", "warn")

    wb = Workbook()
    ws = wb.active
    ws.title = "키워드_성취기준"
    # 사람이 손대는 시트에는 **교육과정 유래 키워드만** 싣는다. 학습된 용어는
    # 가중치·근거 문항수를 달고 있어서 엑셀 한 칸에 담으면 그 숫자가 왕복 중에
    # 사라진다 — 사라진 채 --import-xlsx 하면 애써 배운 사전이 조용히 지워진다.
    # 학습분은 아래 '학습된_용어' 시트에 읽기 전용으로 따로 보여준다.
    n_learned = 0
    for col, (rev, code) in enumerate(columns, start=1):
        ws.cell(row=1, column=col, value=rev)
        ws.cell(row=2, column=col, value=code)
        info_row = (meta_by_rev.get(rev, {}).get(code) or {})
        ws.cell(row=3, column=col, value=info_row.get("unit") or info_row.get("title") or "")
        entry = book.entry(rev, code)
        n_learned += len(entry.get("learned") or [])
        for row_i, keyword in enumerate(entry.get("curriculum") or [], start=XLSX_FIRST_TERM_ROW):
            ws.cell(row=row_i, column=col, value=keyword)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    if n_learned:
        lw = wb.create_sheet("학습된_용어")
        for i, head in enumerate(["개정", "성취기준", "용어", "가중치", "근거 문항수", "로그오즈비"], start=1):
            lw.cell(row=1, column=i, value=head)
        r = 2
        for rev, code in columns:
            for row in book.entry(rev, code).get("learned") or []:
                lw.cell(row=r, column=1, value=rev)
                lw.cell(row=r, column=2, value=code)
                lw.cell(row=r, column=3, value=row.get("term"))
                lw.cell(row=r, column=4, value=row.get("weight"))
                lw.cell(row=r, column=5, value=row.get("df"))
                lw.cell(row=r, column=6, value=row.get("lor"))
                r += 1
        lw.column_dimensions["B"].width = 16
        lw.column_dimensions["C"].width = 30

    info = wb.create_sheet("사용_안내")
    guide = [
        "1행 = 개정 교육과정 연도(2015/2022). 고치지 마세요 — 같은 성취기준 코드를 "
        "두 개정이 함께 쓰는 과목이 있어서, 이 행이 없으면 어느 교육과정의 키워드인지 알 수 없습니다.",
        "2행 = 성취기준 코드(고치지 마세요). 3행 = 단원/설명(참고용, 고쳐도 무시됩니다).",
        f"{XLSX_FIRST_TERM_ROW}행부터 아래로 키워드를 한 줄에 하나씩 입력하세요. 빈 셀은 무시됩니다.",
        "'학습된_용어' 시트는 gw classify --learn 이 라벨된 문항에서 자동으로 캔 것입니다. "
        "읽기 전용이라 여기서 고쳐도 반영되지 않고, --import-xlsx 로도 지워지지 않습니다.",
        f"저장 후: python scripts/gw.py classify --subject {subject.slug} --import-xlsx <이 파일 경로>",
        "성취기준 판정은 동사(이해한다/설명한다)가 아니라 명사(대상) 기준입니다 — "
        "키워드도 현상·물질·개념 같은 '대상' 위주로 넣는 편이 잘 맞습니다.",
    ]
    for i, line in enumerate(guide, start=1):
        info.cell(row=i, column=1, value=line)
    info.column_dimensions["A"].width = 100

    out_path = Path(args.export_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not args.dry_run:
        wb.save(out_path)

    report.count(codes=len(columns),
                 **{f"codes_{rev}": sum(1 for r, _c in columns if r == rev) for rev in revisions})
    report.artifact(str(out_path))
    report.next = f"편집 후: python scripts/gw.py classify --subject {subject.slug} --import-xlsx {out_path}"
    return report.finish()


def _import_xlsx(args) -> int:
    subject = load_subject(args.subject)
    space = Space(subject.slug, getattr(args, "workspace", None)).ensure()
    report = Report("classify_import_xlsx", subject.slug, space)

    try:
        from openpyxl import load_workbook
    except ImportError:
        report.note("openpyxl", "openpyxl 이 설치되어 있지 않다", "error")
        return report.finish()

    in_path = Path(args.import_xlsx)
    if not in_path.exists():
        report.note("import", f"파일이 없다: {in_path}", "error")
        return report.finish()

    wb = load_workbook(in_path, data_only=True)
    ws = wb["키워드_성취기준"] if "키워드_성취기준" in wb.sheetnames else wb.worksheets[0]

    book, _book_path = _keyword_book(subject, space, report)

    # ── 판형을 **모양으로** 알아본다 (스키마 이름을 믿지 않는 것과 같은 이유) ──
    # 새 판형: 1행 개정 / 2행 코드 / 3행 단원 / 4행부터 키워드
    # 옛 판형: 1행 코드 / 2행 단원 / 3행부터 키워드 — 개정 정보가 없다.
    # 옛 파일로 되돌리는 사람이 있을 수 있으므로 둘 다 받는다. 옛 판형이면 개정을
    # 접두사로 역추정하고, 추정했다는 사실을 리포트에 남긴다.
    head = str(ws.cell(row=1, column=1).value or "").strip()
    legacy_sheet = not keywordsio.REVISION_RE.match(head)
    code_row = 1 if legacy_sheet else 2
    first_term_row = 3 if legacy_sheet else XLSX_FIRST_TERM_ROW
    owners = keywordsio.code_owners(tuple(sorted(subject.standard_prefixes or {})
                                          or DEFAULT_REVISIONS))

    codes_seen = 0
    kept_learned = 0
    guessed: list[str] = []
    for col in range(1, ws.max_column + 1):
        code = ws.cell(row=code_row, column=col).value
        if code is None or not str(code).strip():
            continue
        code = str(code).strip()
        if legacy_sheet:
            rev, why = keywordsio.infer_revision(code, subject.standard_prefixes or {}, owners)
            if why:
                guessed.append(code)
        else:
            rev = str(ws.cell(row=1, column=col).value or "").strip()
            if not keywordsio.REVISION_RE.match(rev):
                report.note(code, f"1행의 개정 값 {rev!r} 이 연도 모양이 아니다 — 이 열은 건너뛴다", "warn")
                continue
        kws = []
        for row_i in range(first_term_row, ws.max_row + 1):
            val = ws.cell(row=row_i, column=col).value
            if val is None:
                continue
            val = str(val).strip()
            if val:
                kws.append(val)
        # 엑셀은 교육과정 칸만 왕복한다. 학습된 용어(가중치를 달고 있는 것)는
        # 엑셀에 실리지 않으므로 여기서 **그대로 보존**해야 한다. 예전처럼 통째로
        # 덮어쓰면 사람이 키워드 하나 고치려고 엑셀을 왕복한 순간 학습 결과가 날아간다.
        learned = book.entry(rev, code).get("learned") or []
        kept_learned += len(learned)
        book.set_entry(rev, code, {"curriculum": kws, "learned": learned})
        codes_seen += 1

    if legacy_sheet:
        report.note(in_path.name,
                    f"1행이 개정이 아니라 코드다 — 옛 판형으로 읽었다"
                    + (f". {len(guessed)}개 코드는 개정을 접두사로도 못 갈라 추정했다"
                       f"({', '.join(sorted(guessed)[:4])}{'…' if len(guessed) > 4 else ''}) — "
                       f"--export-xlsx 로 새로 내보내 쓰는 편이 안전하다" if guessed else ""),
                    "warn" if guessed else "info")

    out_path = keywordsio.write_target(space, subject.keywords_path,
                                       bool(getattr(args, "install", False)))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    keywordsio.save(out_path, book, dry_run=bool(args.dry_run))

    report.count(codes_updated=codes_seen,
                 codes_total=sum(len(book.revision(r)) for r in book.known_revisions()),
                 learned_kept=kept_learned)
    report.artifact(str(out_path))
    if not getattr(args, "install", False):
        report.note("keywords.json",
                    keywordsio.install_hint(space, subject.keywords_path,
                                            f"classify --subject {subject.slug} --import-xlsx ..."),
                    "info")
    report.next = f"python scripts/gw.py classify --subject {subject.slug}"
    return report.finish()


# ---------------------------------------------------------------- CLI
def register(parser) -> None:
    parser.add_argument("--subject", required=True, help="과목 슬러그")
    parser.add_argument("--revision", choices=["2015", "2022", "both"], default="both",
                         help="분류 대상 개정 교육과정 (기본: 정의된 것 모두)")
    parser.add_argument("--apply", metavar="FILE",
                         help="LLM 판정 결과(classify_result.json)를 items/ 에 반영")
    parser.add_argument("--learn", action="store_true",
                         help="라벨된 문항(by=manual/llm)에서 변별 용어를 배워 keywords.json 에 병합한다")
    parser.add_argument("--calibrate", action="store_true",
                         help="라벨(by=manual/llm)로 자기 정확도를 재고 권장 임계값을 "
                              "calibration.json 에 쓴다. 표본이 llm 뿐이면 '자기 채점'으로 표시한다")
    parser.add_argument("--export-xlsx", metavar="PATH", help="keywords.json 을 엑셀로 내보내기")
    parser.add_argument("--import-xlsx", metavar="PATH", help="엑셀을 keywords.json 으로 되돌리기")
    parser.add_argument("--dry-run", action="store_true", help="파일을 쓰지 않고 리포트만 미리 본다")
    parser.add_argument("--force", action="store_true",
                         help="이미 분류된 항목도 다시 계산한다 (by=manual 인 값은 절대 덮지 않는다)")
    parser.add_argument("--only", metavar="QID|EXAM_ID,...",
                         help="이 문항/회차만 처리. --learn 에서는 학습 대상을, "
                              "--calibrate 에서는 채점 대상을 제한한다 (학습/평가 분리에 쓴다)")
    parser.add_argument("--quiet", action="store_true", help="stdout 요약도 생략한다")
    parser.add_argument("--workspace", help="작업 공간 경로 직접 지정 (기본 workspace/<slug>)")
    parser.add_argument("--install", action="store_true",
                         help="산출물(keywords.json·calibration.json)을 작업공간이 아니라 "
                              "subjects/<slug>/ 에 쓴다. 기본은 작업공간이다 — 저장소에 "
                              "파일이 새로 생기면 git pull 이 막힌다(docs 참조)")


def run(args) -> int:
    # 부기능(분류/학습/보정/적용/엑셀 왕복)은 상호 배타적이다 — gw.py 계약상 명령
    # 하나가 여러 산출물을 만들 수는 있어도, 이들은 서로 다른 입력을 요구해서
    # 한 번의 실행에 섞으면 어느 리포트를 신뢰해야 할지 모호해진다.
    # 특히 --learn 과 --calibrate 를 한 번에 돌리면 방금 배운 문항으로 자기를
    # 채점하게 되므로, 순서를 사람이 정하도록 일부러 갈라 둔다.
    try:
        if args.apply:
            return _apply(args)
        if getattr(args, "learn", False) and getattr(args, "calibrate", False):
            print("[FAIL] classify --learn 과 --calibrate 는 함께 쓸 수 없다. "
                  "--learn 을 먼저 돌리고, 학습에 쓰지 않은 회차로 --calibrate 를 돌려라.")
            return 1
        if getattr(args, "learn", False):
            return _learn(args)
        if getattr(args, "calibrate", False):
            return _calibrate(args)
        if args.export_xlsx:
            return _export_xlsx(args)
        if args.import_xlsx:
            return _import_xlsx(args)
        return _classify(args)
    except FileNotFoundError as exc:
        # load_subject() 가 슬러그를 못 찾으면 파이썬 트레이스백을 던진다 — 이건
        # common/subjects.py 소관이라 고치지 않지만, 여기서라도 잡아 CONTRACT 5절이
        # 요구하는 "stdout 에는 리포트 경로와 한 줄 요약만" 규칙에 맞춰 조용히
        # 실패한다. 이 시점엔 subject 를 못 읽었으니 리포트 파일을 남길 workspace
        # 조차 확정할 수 없어(오타 난 슬러그로 workspace 폴더를 새로 만들면 그것도
        # 지저분하다) 리포트 없이 한 줄만 찍는다.
        print(f"[FAIL] classify {exc}")
        return 1
